#!/usr/bin/env python3
"""
Weekly Content Pipeline Orchestrator

Multi-client: reads keywords, brand terms, and output paths from client config.

Called in stages by .claude/commands/weekly-pipeline.md:

  Stage 1 — scrape topics:
    python scripts/weekly_pipeline.py --action scrape --week-of 2026-02-16 --count 100 --output /tmp/weekly_topics.json

  Stage 2 — create workbook:
    python scripts/weekly_pipeline.py --action create-workbook --week-of 2026-02-16

  Stage 3 — save one content item (called once per topic by Claude):
    python scripts/weekly_pipeline.py --action save-content \
      --week-of 2026-02-16 \
      --content-file /tmp/weekly_content_1.json

  Stage 4 — finalize (macOS notification + summary):
    python scripts/weekly_pipeline.py --action finalize --week-of 2026-02-16

  Stage 5 — sync to Airtable (optional, if airtable.enabled in client config):
    python scripts/weekly_pipeline.py --action sync-airtable --week-of 2026-02-16

Content JSON format for save-content:
{
  "date": "2026-02-16",
  "day": "Mon",
  "topic": "Topic text here",
  "platform": "Twitter",
  "format": "thread",
  "content": "English content...",
  "image_prompt": "Detailed prompt for nano_banana.py...",
  "image_path": "outputs/content/bobe/images/2026-02-16-weekly/2026-02-16_mon_topic-slug_twitter.png",
  "hashtags": ["#DeFi", "#TradingBot"],
  "content_ru": "Russian content...",
  "image_prompt_ru": "Russian image prompt with Cyrillic headline...",
  "image_path_ru": "outputs/content/bobe/images/2026-02-16-weekly/2026-02-16_mon_topic-slug_twitter_ru.png",
  "hashtags_ru": ["#DeFi", "#Крипто"],
  "status": "Draft"
}
"""

import os
import sys
import json
import argparse
import subprocess
from datetime import datetime, timedelta
from pathlib import Path

# Make scripts/ importable as a package
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# Import styling helpers from excel_manager
from excel_manager import (
    style_header_cell, style_data_cell,
    COLOR_DARK_BG, COLOR_BLUE, COLOR_GREEN, COLOR_WHITE, COLOR_LIGHT_ROW,
)

import client_config

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
POSTS_PER_DAY = 3


def _get_output_dir(client_id=None):
    """Get the client-scoped output directory."""
    return client_config.get_output_dir(client_id)


def get_week_of(date_str=None):
    """Return Monday of the current (or specified) week as YYYY-MM-DD."""
    if date_str:
        d = datetime.strptime(date_str, "%Y-%m-%d")
    else:
        d = datetime.now()
    monday = d - timedelta(days=d.weekday())
    return monday.strftime("%Y-%m-%d")


def scrape_weekly_topics(week_of, count=100, mock=False, output_path=None, client_id=None):
    """
    Scrape Twitter for the past 7 days using apify_scraper.
    Falls back gracefully if scraping fails.
    Returns ranked list of topic dicts.
    """
    from apify_scraper import (
        scrape_twitter, filter_by_relevance, rank_by_engagement, mock_scrape
    )

    keywords = client_config.get_keywords(client_id)

    if mock:
        print("Running in MOCK mode, no API calls.")
        posts = []
        for _ in range(5):
            posts += mock_scrape("twitter")
        filtered = filter_by_relevance(posts, keywords, client_id=client_id)
        ranked = rank_by_engagement(filtered)
        if output_path:
            with open(output_path, "w") as f:
                json.dump(ranked, f, indent=2, default=str)
        print(f"Mock: {len(ranked)} topics generated.")
        return ranked

    since_date = (datetime.strptime(week_of, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")
    print(f"Scraping Twitter for past 7 days (since {since_date})...")

    all_posts = []
    try:
        tweets = scrape_twitter(keywords, count=count, since_date=since_date)
        print(f"  Got {len(tweets)} tweets")
        all_posts.extend(tweets)
    except Exception as e:
        print(f"  Twitter scraping failed: {e}")

    filtered = filter_by_relevance(all_posts, keywords, client_id=client_id)
    ranked = rank_by_engagement(filtered)
    print(f"  {len(ranked)} relevant posts after filtering")

    if output_path:
        with open(output_path, "w") as f:
            json.dump(ranked, f, indent=2, default=str)
        print(f"  Saved to: {output_path}")

    return ranked


def create_weekly_workbook(week_of, client_id=None):
    """
    Create outputs/content/{client_id}/{week_of}-weekly-content.xlsx with:
    - Sheet 1: Topics (same columns as daily, Week Of instead of Date)
    - Sheet 2: Content (adds Day column as column B)
    """
    output_dir = _get_output_dir(client_id)
    wb = Workbook()

    # --- Sheet 1: Topics ---
    ws1 = wb.active
    ws1.title = "Topics"

    topic_headers = [
        "Week Of", "Platform", "Topic Summary", "Original Text",
        "Author", "URL", "Engagement Score", "Relevance Score", "Source Type",
    ]
    topic_col_widths = [12, 10, 35, 60, 20, 40, 16, 16, 14]

    for i, (header, width) in enumerate(zip(topic_headers, topic_col_widths), 1):
        cell = ws1.cell(row=1, column=i, value=header)
        style_header_cell(cell)
        ws1.column_dimensions[get_column_letter(i)].width = width

    ws1.row_dimensions[1].height = 35
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Content (with Day column) ---
    ws2 = wb.create_sheet("Content")

    content_headers = [
        "Date", "Day", "Topic", "Platform Target", "Format",
        "Content", "Image Prompt", "Image Path", "Hashtags",
        "Content_RU", "Image_Prompt_RU", "Image_Path_RU", "Hashtags_RU", "Status",
    ]
    content_col_widths = [12, 6, 30, 14, 10, 70, 50, 40, 35, 70, 50, 40, 35, 12]

    for i, (header, width) in enumerate(zip(content_headers, content_col_widths), 1):
        cell = ws2.cell(row=1, column=i, value=header)
        style_header_cell(cell)
        ws2.column_dimensions[get_column_letter(i)].width = width

    ws2.row_dimensions[1].height = 35
    ws2.freeze_panes = "A2"

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{week_of}-weekly-content.xlsx"
    wb.save(str(path))
    return str(path)


def append_content_row(week_of, content_item, client_id=None):
    """
    Append one content row to the weekly workbook's Content sheet.
    content_item must have: date, day, topic, platform, format, content,
                            image_prompt, image_path, hashtags, status
    """
    output_dir = _get_output_dir(client_id)
    path = output_dir / f"{week_of}-weekly-content.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"Weekly workbook not found: {path}. Run --action create-workbook first.")

    wb = load_workbook(str(path))
    ws = wb["Content"]
    next_row = ws.max_row + 1
    row_idx = next_row - 2  # for alternating color (0-based after header)

    hashtags = content_item.get("hashtags", [])
    hashtag_str = ", ".join(hashtags) if isinstance(hashtags, list) else str(hashtags)

    hashtags_ru = content_item.get("hashtags_ru", [])
    hashtag_ru_str = ", ".join(hashtags_ru) if isinstance(hashtags_ru, list) else str(hashtags_ru)

    values = [
        content_item.get("date", ""),
        content_item.get("day", ""),
        content_item.get("topic", ""),
        content_item.get("platform", ""),
        content_item.get("format", ""),
        content_item.get("content", ""),
        content_item.get("image_prompt", ""),
        content_item.get("image_path", ""),
        hashtag_str,
        content_item.get("content_ru", ""),
        content_item.get("image_prompt_ru", ""),
        content_item.get("image_path_ru", ""),
        hashtag_ru_str,
        content_item.get("status", "Draft"),
    ]

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        style_data_cell(cell, row_idx)

    ws.row_dimensions[next_row].height = 80
    wb.save(str(path))


def translate_text_to_russian(en_content, platform="twitter", client_id=None):
    """
    Translate English content to Russian using Google Gemini API.
    Preserves --- tweet separators, keeps English hashtags, adds Cyrillic hashtags.
    Returns translated text string.
    """
    from google import genai

    api_key = os.getenv("GOOGLE_AI_API_KEY")
    if not api_key:
        raise RuntimeError("GOOGLE_AI_API_KEY not set in .env")

    config = client_config.load_config(client_id)
    brand_terms = ", ".join(config.get("content", {}).get("brand_terms_keep", []))
    tone = config.get("content", {}).get("tone", "professional and educational")

    client = genai.Client(api_key=api_key)

    platform_note = ""
    if platform == "twitter":
        platform_note = (
            "This is a Twitter thread. Each tweet is separated by '---' on its own line. "
            "Keep the '---' separators exactly as they are. "
            "Keep each tweet under 280 characters. "
        )
    elif platform == "telegram":
        platform_note = "This is a Telegram post. Preserve formatting and line breaks. "

    prompt = f"""Translate the following English crypto/fintech content to Russian.

Rules:
- Translate naturally, not word-for-word. Use professional Russian crypto/fintech terminology.
- NEVER use em-dashes, en-dashes, or double-hyphens as punctuation. Use commas, colons, or rephrase instead.
- {platform_note}
- Keep brand names unchanged: {brand_terms}
- Keep any hashtags at the end unchanged (English hashtags stay as-is).
- The tone should be {tone}.

Content to translate:
{en_content}"""

    response = client.models.generate_content(
        model="gemini-2.0-flash",
        contents=prompt,
    )
    return response.text.strip()


def translate_hashtags_to_russian(en_hashtags, client_id=None):
    """
    Translate English hashtags to Russian: keep originals, add Cyrillic versions.
    en_hashtags: comma-separated string like "#DeFi, #TradingBot"
    Returns comma-separated string with both EN and RU hashtags.
    """
    from google import genai

    api_key = os.getenv("GOOGLE_AI_API_KEY")
    if not api_key:
        raise RuntimeError("GOOGLE_AI_API_KEY not set in .env")

    client = genai.Client(api_key=api_key)

    prompt = f"""Given these English crypto hashtags: {en_hashtags}

Create a combined hashtag list that includes:
1. The original English hashtags (keep them exactly)
2. Russian/Cyrillic equivalents for each (where translation makes sense)

Return ONLY a comma-separated list of hashtags, nothing else.
Example input: #DeFi, #TradingBot
Example output: #DeFi, #TradingBot, #Крипто, #ТорговыйБот"""

    response = client.models.generate_content(
        model="gemini-2.0-flash",
        contents=prompt,
    )
    return response.text.strip()


def update_ru_columns(week_of, row_index, content_ru, image_prompt_ru, image_path_ru, hashtags_ru, client_id=None):
    """
    Update columns J-M (Content_RU, Image_Prompt_RU, Image_Path_RU, Hashtags_RU)
    for a specific row in the workbook.
    row_index: 1-based index into Content sheet data rows (1 = first data row = Excel row 2).
    """
    output_dir = _get_output_dir(client_id)
    path = output_dir / f"{week_of}-weekly-content.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"Weekly workbook not found: {path}")

    wb = load_workbook(str(path))
    ws = wb["Content"]

    excel_row = row_index + 1  # +1 for header row

    # Columns J=10, K=11, L=12, M=13
    ws.cell(row=excel_row, column=10, value=content_ru)
    ws.cell(row=excel_row, column=11, value=image_prompt_ru)
    ws.cell(row=excel_row, column=12, value=image_path_ru)
    ws.cell(row=excel_row, column=13, value=hashtags_ru)

    wb.save(str(path))


def finalize(week_of, client_id=None):
    """Send macOS notification and print completion summary."""
    output_dir = _get_output_dir(client_id)
    config = client_config.load_config(client_id)
    display_name = config.get("display_name", client_id or "Client")

    path = output_dir / f"{week_of}-weekly-content.xlsx"
    row_count = 0

    if path.exists():
        wb = load_workbook(str(path), read_only=True, data_only=True)
        if "Content" in wb.sheetnames:
            row_count = wb["Content"].max_row - 1  # subtract header
        wb.close()

    msg = f"{row_count} content items ready in {week_of}-weekly-content.xlsx"
    title = f"{display_name} Weekly Pipeline Complete"

    # macOS notification via osascript
    script = f'display notification "{msg}" with title "{title}"'
    subprocess.run(["osascript", "-e", script], capture_output=True)

    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"  Week of: {week_of}")
    print(f"  {row_count} content rows saved")
    print(f"  File: {path}")
    print(f"  View: /view-content week:{week_of}")
    print(f"{'='*60}\n")




def regenerate_topic_content(xlsx_path, topic_index, client_id=None, mock=False):
    """
    Regenerate EN Twitter + Telegram content for a specific topic using Gemini.

    Finds the two rows (Twitter + Telegram) for topic_index in the Content sheet,
    regenerates content via Gemini using the client's content guidelines,
    writes the new content back to column F (Content), and saves the workbook.

    Returns: {"twitter": "...", "telegram": "..."}
    """
    from google import genai

    api_key = os.getenv("GOOGLE_AI_API_KEY")
    if not api_key:
        raise RuntimeError("GOOGLE_AI_API_KEY not set in .env")

    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    # Load config and guidelines
    config = client_config.load_config(client_id)
    display_name = config.get("display_name", client_id or "Brand")
    tone = config.get("content", {}).get("tone", "transparent, educational")
    voice = config.get("content", {}).get("voice", "")
    pillars = "; ".join(config.get("content", {}).get("messaging_pillars", []))
    ctas = "; ".join(config.get("content", {}).get("cta_examples", []))
    hashtags = ", ".join(config.get("content", {}).get("hashtags", []))

    guidelines_path = client_config.get_content_guidelines_path(client_id)
    guidelines_snippet = ""
    if guidelines_path.exists():
        guidelines_snippet = guidelines_path.read_text(encoding="utf-8")[:1500]

    # Read workbook to find topic rows in order
    wb_read = load_workbook(str(xlsx), read_only=True, data_only=True)
    ws_read = wb_read["Content"]

    topic_order = []
    topic_rows = {}  # topic_name -> list of {excel_row, platform, format, date, day}
    for row_idx, row in enumerate(ws_read.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[2]:
            continue
        tname = row[2]
        if tname not in topic_rows:
            topic_order.append(tname)
            topic_rows[tname] = []
        topic_rows[tname].append({
            "excel_row": row_idx,
            "platform": (row[3] or "").lower(),
            "format": row[4] or "thread",
            "date": str(row[0]) if row[0] else "",
            "day": row[1] or "",
        })
    wb_read.close()

    if topic_index >= len(topic_order):
        raise ValueError(f"topic_index {topic_index} out of range (max {len(topic_order)-1})")

    topic_name = topic_order[topic_index]
    rows_for_topic = topic_rows[topic_name]

    if not rows_for_topic:
        raise RuntimeError(f"No rows found for topic: {topic_name}")

    first_row = rows_for_topic[0]
    day = first_row["day"]
    date = first_row["date"]

    if mock:
        new_twitter = f"[Mock regenerated Twitter] {topic_name[:60]}"
        new_telegram = f"[Mock regenerated Telegram] {topic_name[:60]}"
    else:
        gc = genai.Client(api_key=api_key)

        def _gen(platform_name, fmt_hint):
            if platform_name == "twitter":
                if fmt_hint in ("thread", ""):
                    platform_rules = (
                        "Write a 5-tweet thread. Separate tweets with exactly ---. "
                        "Tweet 1: hook or bold claim. Tweets 2-3: education or insight. "
                        f"Tweet 4: how {display_name} connects. "
                        "Tweet 5: soft CTA. Each tweet must be 280 chars or fewer."
                    )
                    format_desc = "Twitter thread (5 tweets separated by ---)"
                else:
                    platform_rules = "Write a single tweet, 280 chars or fewer. Hook + insight + 2-3 hashtags."
                    format_desc = "Single tweet"
            else:
                platform_rules = (
                    "Write a Telegram post between 400 and 1200 characters. "
                    "Educational tone, explain the why. Use line breaks generously. "
                    "End with an engagement question to the reader."
                )
                format_desc = "Telegram post"

            prompt = f"""You are a content writer for {display_name}.

Brand voice: {tone}
Voice: {voice}
Messaging pillars: {pillars}
CTAs: {ctas}
Hashtags: {hashtags}

CRITICAL RULES:
- NEVER use em-dash, en-dash, or double-hyphen as punctuation
- Replace with commas, colons, or rephrase entirely
- The --- tweet separator is the ONLY exception
- No guaranteed return claims, no hype

Content guidelines:
{guidelines_snippet[:800]}

Topic: "{topic_name}"
Day: {day}, Date: {date}
Platform: {platform_name.capitalize()}
Format: {format_desc}

{platform_rules}

Return ONLY the content text, no JSON, no extra explanation."""

            resp = gc.models.generate_content(model="gemini-2.0-flash", contents=prompt)
            return resp.text.strip()

        # Determine twitter format from existing rows
        twitter_fmt = "thread"
        for r in rows_for_topic:
            if "twitter" in r["platform"]:
                twitter_fmt = r["format"] or "thread"
                break

        new_twitter = _gen("twitter", twitter_fmt)
        new_telegram = _gen("telegram", "long-form")

    # Write back to workbook
    wb_write = load_workbook(str(xlsx))
    ws_write = wb_write["Content"]

    for row_info in rows_for_topic:
        excel_row = row_info["excel_row"]
        platform = row_info["platform"]
        if "twitter" in platform:
            ws_write.cell(row=excel_row, column=6, value=new_twitter)
        elif "telegram" in platform:
            ws_write.cell(row=excel_row, column=6, value=new_telegram)

    wb_write.save(str(xlsx))
    return {"twitter": new_twitter, "telegram": new_telegram}

def main():
    parser = argparse.ArgumentParser(description="Weekly Pipeline Orchestrator")
    parser.add_argument("--action",
                        choices=["scrape", "create-workbook", "save-content", "finalize", "sync-airtable"],
                        required=True,
                        help="Pipeline stage to execute")
    parser.add_argument("--week-of", default=None,
                        help="Monday of target week (YYYY-MM-DD). Defaults to current week's Monday.")
    parser.add_argument("--count", type=int, default=100,
                        help="Max posts to scrape (--action scrape only)")
    parser.add_argument("--output", default=None,
                        help="Output JSON path (--action scrape only)")
    parser.add_argument("--content-file", default=None,
                        help="Path to content JSON file (--action save-content only)")
    parser.add_argument("--mock", action="store_true",
                        help="Use mock data instead of real API calls")
    client_config.add_client_arg(parser)

    args = parser.parse_args()
    active_client = client_config.resolve_client(args)
    week_of = get_week_of(args.week_of) if args.week_of else get_week_of()

    if args.action == "scrape":
        scrape_weekly_topics(week_of, count=args.count, mock=args.mock, output_path=args.output, client_id=active_client)

    elif args.action == "create-workbook":
        path = create_weekly_workbook(week_of, client_id=active_client)
        print(f"Weekly workbook created: {path}")

    elif args.action == "save-content":
        if not args.content_file:
            print("Error: --content-file is required for save-content action")
            sys.exit(1)
        with open(args.content_file) as f:
            item = json.load(f)
        append_content_row(week_of, item, client_id=active_client)
        print(f"Saved: [{item.get('day', '?')}] {item.get('topic', '')[:60]} ({item.get('platform', '')})")

    elif args.action == "finalize":
        finalize(week_of, client_id=active_client)

    elif args.action == "sync-airtable":
        cmd = [sys.executable, str(Path(__file__).parent / "airtable_sync.py"),
               "--week-of", week_of]
        if active_client:
            cmd += ["--client", active_client]
        if args.mock:
            cmd.append("--mock")
        result = subprocess.run(cmd, capture_output=False)
        sys.exit(result.returncode)


if __name__ == "__main__":
    main()
