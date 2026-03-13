#!/usr/bin/env python3
"""
Standalone Pipeline Runner

End-to-end pipeline: scrape topics → generate content via Gemini API →
images via WaveSpeed → Excel workbook → Airtable sync → build static site.

Runs without Claude Code. Used by GitHub Actions for remote pipeline execution.
The Claude-driven /weekly-pipeline command remains the gold standard for
quality-critical runs (better brand voice consistency).

Usage:
  python scripts/pipeline_runner.py --client bobe
  python scripts/pipeline_runner.py --client bobe --week-of 2026-02-23
  python scripts/pipeline_runner.py --mock
  python scripts/pipeline_runner.py --skip-images --skip-airtable --skip-deploy

Flags:
  --client         Client ID (default: reads .active-client or bobe)
  --week-of        Week start date YYYY-MM-DD (default: Monday of current week)
  --mock           Dry run, no API calls
  --skip-images    Skip image generation
  --skip-airtable  Skip Airtable sync
  --skip-deploy    Skip static site build (GH Actions handles deploy separately)
"""

import os
import sys
import json
import re
import time
import logging
import subprocess
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    level=logging.INFO,
)
logger = logging.getLogger("pipeline")

from dotenv import load_dotenv
load_dotenv()

sys.path.insert(0, str(Path(__file__).parent))
import client_config
import bucket_generators
from client_config import get_api_key
from utils import is_cyrillic

try:
    import airtable_writer
    import r2_uploader
    HAS_CLOUD_MODULES = True
except ImportError:
    HAS_CLOUD_MODULES = False

# Use the same Python interpreter that launched this script (works on macOS + Linux + GH Actions)
# Quoted to handle paths with spaces (e.g. "Claude Code" directory on macOS)
PY = f'"{sys.executable}"'

PROJECT_ROOT = Path(__file__).parent.parent
DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


# ── Utilities ─────────────────────────────────────────────────────────────────

def get_monday(date_str=None):
    """Return Monday of current (or specified) week as YYYY-MM-DD."""
    if date_str:
        d = datetime.strptime(date_str, "%Y-%m-%d")
    else:
        d = datetime.now()
    monday = d - timedelta(days=d.weekday())
    return monday.strftime("%Y-%m-%d")


def topic_slug(topic, max_chars=30):
    """Convert topic text to a safe filename slug."""
    s = topic.lower()[:max_chars]
    s = re.sub(r"[^a-z0-9\s]", "", s)
    s = re.sub(r"\s+", "_", s.strip())
    return s or "topic"


def make_image_path(client_id, week_of, day, slug, platform, ru=False):
    """Build the expected image output path."""
    suffix = "_ru" if ru else ""
    return (
        f"outputs/content/{client_id}/images/{week_of}-weekly/"
        f"{week_of}_{day.lower()}_{slug}_{platform.lower()}{suffix}.png"
    )


def run_cmd(cmd, check=True):
    """Run a shell command, raise on non-zero exit if check=True."""
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    if result.stdout.strip():
        print(f"    {result.stdout.strip()}")
    if result.returncode != 0:
        if result.stderr.strip():
            print(f"    stderr: {result.stderr.strip()[:300]}")
        if check:
            raise RuntimeError(
                f"Command failed (exit {result.returncode}): {cmd[:120]}"
            )
    return result


# ── Gemini API ─────────────────────────────────────────────────────────────────

def call_gemini(prompt, model="gemini-2.0-flash", client_id=None):
    """Call Gemini API and return the text response."""
    api_key = get_api_key(client_id or "bobe", "gemini")
    if not api_key:
        raise ValueError("GOOGLE_AI_API_KEY not set. Check your .env file.")
    try:
        from google import genai
        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(model=model, contents=prompt)
        return response.text
    except ImportError:
        try:
            import google.generativeai as genai
            genai.configure(api_key=api_key)
            m = genai.GenerativeModel(model)
            response = m.generate_content(prompt)
            return response.text
        except ImportError:
            raise ImportError(
                "No Gemini SDK found. Install: pip install google-genai"
            )


def extract_json(text):
    """Extract the first JSON array or object from a text response."""
    # Try fenced code block first
    match = re.search(r"```(?:json)?\s*([\[\{][\s\S]*?[\]\}])\s*```", text)
    if match:
        return json.loads(match.group(1))
    # Try raw JSON
    match = re.search(r"([\[\{][\s\S]*[\]\}])", text)
    if match:
        return json.loads(match.group(1))
    raise ValueError(f"No valid JSON found in response:\n{text[:400]}")


# ── Prompts ────────────────────────────────────────────────────────────────────

TOPIC_POOL_PROMPT = """\
You are a content strategist for {display_name}.

Brand: {display_name} — {tagline}
Tone: {tone}
Target audience: retail crypto users age 20-35, $500-$20k in crypto, exhausted from emotional trading
Messaging pillars: {pillars}
Keywords: {keywords}

Scraped live topics from Twitter this week ({n_scraped} found):
{scraped_summary}

Generate a weekly schedule of exactly 21 content topics (3 per day, 7 days).
Use scraped live topics where directly relevant to the brand; fill the rest with evergreen topics.
Live/trending topics go to Mon-Wed priority; evergreen fills Thu-Sun and remaining slots.
Each day: mix of Pain Point, Education, and Transparency/Product angles.
Vary angles so consecutive days do not repeat the same angle.

CRITICAL RULES:
- NEVER use em-dashes, en-dashes, or double-hyphens (-- or — or –) in topic text
- Replace with commas, colons, or rephrase
- Topics must be concise (max 80 chars) and specific
- No hype, no guaranteed return claims

Return ONLY a JSON array of exactly 21 objects, no other text:
[
  {{
    "topic_num": 1,
    "day": "Mon",
    "date": "{date_mon}",
    "topic": "topic text here",
    "angle": "Pain Point",
    "source": "Live"
  }},
  ...
]

Day/date schedule: Mon {date_mon}, Tue {date_tue}, Wed {date_wed}, Thu {date_thu}, Fri {date_fri}, Sat {date_sat}, Sun {date_sun}
3 topics per day, numbered 1-21. Angles rotate: Pain Point, Education, Transparency/Product."""


CONTENT_GEN_PROMPT = """\
You are a content writer for {display_name}.

Brand voice: {tone}
Voice: {voice}
Product: {tagline}
Messaging pillars: {pillars}
CTAs: {ctas}
Hashtags: {hashtags}

CRITICAL RULES:
- NEVER use — (em-dash), – (en-dash), or -- (double-hyphen) as punctuation
- Replace with commas, colons, or rephrase entirely
- The --- tweet separator is the ONLY exception (structural separator between tweets)
- No guaranteed return claims, no hype, no moon/rocket emojis

Topic #{topic_num}: "{topic}"
Angle: {angle}
Day: {day}, Date: {date}
Platform: {platform}
Format: {format_desc}

{source_context}{platform_rules}

Generate content in BOTH English and Russian.

Russian rules:
- Conversational tone, like a knowledgeable friend. Not formal, not slang.
- Twitter: ~280 char target (allow up to 340 since Russian runs ~30% longer)
- Telegram: maintain structure and engagement question at end
- Keep English hashtags (#DeFi, #USDT, #BoBe etc.) AND add 1-2 Cyrillic ones (#Крипто, #АвтоТрейдинг, #Доходность)
- The --- tweet separator stays unchanged in Russian threads
- No em-dashes, en-dashes, or double-hyphens in Russian either

Image prompt guidelines:
- Include brand mascot: {mascot}
- Background style: {bg_style}
- Image style preset: {style_preset}
- Include a bold prominent headline in the image text

Return ONLY a JSON object, no other text:
{{
  "content": "English content here",
  "image_prompt": "Detailed image prompt with mascot, background, style, and bold English headline",
  "hashtags": ["#Tag1", "#Tag2", "#Tag3"],
  "content_ru": "Russian content here",
  "image_prompt_ru": "Same as image_prompt but with bold RUSSIAN/CYRILLIC headline instead of English",
  "hashtags_ru": ["#Tag1", "#Tag2", "#Крипто"]
}}"""


def get_platform_rules(config, platform, fmt):
    """Return platform-specific content rules string."""
    pf = config.get("content", {}).get("platform_formats", {})
    if platform.lower() == "twitter":
        if fmt == "thread":
            n = pf.get("twitter", {}).get("thread_tweets", 5)
            return (
                f"Write a {n}-tweet thread. Separate tweets with exactly ---. "
                f"Tweet 1: hook or bold claim. Tweets 2-3: education or insight. "
                f"Tweet 4: how {config.get('display_name', 'the brand')} connects. "
                f"Tweet 5: soft CTA. Each tweet must be 280 chars or fewer."
            )
        else:
            return "Write a single tweet, 280 chars or fewer. Hook + insight + 2-3 hashtags."
    elif platform.lower() == "telegram":
        min_c = pf.get("telegram", {}).get("min_chars", 400)
        max_c = pf.get("telegram", {}).get("max_chars", 1200)
        return (
            f"Write a Telegram post between {min_c} and {max_c} characters. "
            f"Educational tone, explain the 'why'. Use line breaks generously. "
            f"End with an engagement question to the reader."
        )
    return "Write content appropriate for this platform in the brand voice."


def get_style_preset(config, angle):
    """Get image style key and description for a given topic angle."""
    style_map = config.get("image", {}).get("angle_style_map", {})
    style_key = style_map.get(angle, "minimal")
    presets = config.get("image", {}).get("style_presets", {})
    style_desc = presets.get(style_key, "clean minimal design on dark background")
    return style_key, style_desc


# ── Mock data ──────────────────────────────────────────────────────────────────

def make_mock_topics(day_dates):
    """Generate mock 21-topic schedule for dry runs."""
    topics = []
    angles = ["Pain Point", "Education", "Product"]
    for i in range(21):
        day = DAYS[i // 3]
        topics.append({
            "topic_num": i + 1,
            "day": day,
            "date": day_dates[day],
            "topic": f"Mock topic {i+1}: crypto automation insight for testing",
            "angle": angles[i % 3],
            "source": "Evergreen",
        })
    return topics


MOCK_CONTENT = {
    "content": "Mock EN content. No real API calls made in --mock mode.",
    "image_prompt": "Mock image prompt: BoBe mascot, dark navy background, bold headline.",
    "hashtags": ["#DeFi", "#TradingBot", "#BoBe"],
    "content_ru": "Мок контент. API вызовы не выполнялись в режиме --mock.",
    "image_prompt_ru": "Мок промпт: маскот BoBe, тёмно-синий фон, жирный заголовок на кириллице.",
    "hashtags_ru": ["#DeFi", "#TradingBot", "#BoBe", "#Крипто"],
}


# ── Main pipeline ──────────────────────────────────────────────────────────────

def run_pipeline(client_id, week_of, mock=False, skip_images=False,
                 skip_airtable=False, skip_deploy=False, export_excel=False,
                 parallel_workers=4):
    """Run the full autonomous content pipeline."""

    config = client_config.load_config(client_id)
    display_name = config.get("display_name", client_id)
    output_dir = client_config.get_output_dir(client_id)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"  {display_name} Pipeline Runner")
    print(f"  Client: {client_id} | Week: {week_of}")
    print(f"  Mock: {mock} | Skip images: {skip_images}")
    print(f"{'='*60}\n")

    # Day dates for the full week
    week_start = datetime.strptime(week_of, "%Y-%m-%d")
    day_dates = {
        DAYS[i]: (week_start + timedelta(days=i)).strftime("%Y-%m-%d")
        for i in range(7)
    }

    # Content config shortcuts
    tone = config.get("content", {}).get("tone", "transparent, educational")
    voice = config.get("content", {}).get("voice", "")
    pillars = "; ".join(config.get("content", {}).get("messaging_pillars", []))
    ctas = "; ".join(config.get("content", {}).get("cta_examples", []))
    hashtags = ", ".join(config.get("content", {}).get("hashtags", []))
    keywords = ", ".join(client_config.get_keywords(client_id)[:15])
    mascot = config.get("brand", {}).get("mascot_description", "brand mascot")
    bg_style = config.get("brand", {}).get("background_style", "dark navy background")

    errors = []

    # ── Phase 1: Scrape Topics (for trending bucket) ──────────────────────────
    print("Phase 1: Scraping trending topics...")
    scraped_topics = []

    if not mock:
        scrape_output = "/tmp/pipeline_runner_scraped.json"
        try:
            run_cmd(
                f"{PY} scripts/apify_scraper.py "
                f"--platform twitter --count 50 --days 7 --top 20 "
                f"--output {scrape_output} --client {client_id}",
                check=False,
            )
            if Path(scrape_output).exists():
                with open(scrape_output) as f:
                    scraped_topics = json.load(f)
                print(f"  Scraped {len(scraped_topics)} topics")
            else:
                print("  Warning: scraper produced no output, using evergreen fallback")
        except Exception as e:
            print(f"  Warning: scraping failed ({e}), using evergreen fallback")
    else:
        print("  [mock] Skipping scrape")

    # ── Phase 2: Assemble 21 Topics (3 buckets × 7 topics) ───────────────────
    print("\nPhase 2: Assembling 3-bucket topic pool...")

    content_types = client_config.get_content_types(client_id)
    bucket_size = client_config.get_bucket_size(client_id)

    # Load any existing client inputs for this week
    inputs_file = output_dir / f"{week_of}-bucket-inputs.json"
    client_inputs = {}
    if inputs_file.exists():
        try:
            client_inputs = json.loads(inputs_file.read_text(encoding="utf-8"))
        except Exception:
            pass

    if mock:
        all_bucket_lists = []
        for bucket_type in content_types:
            bucket_topics = bucket_generators.generate_bucket(
                bucket_type, config, week_of, day_dates,
                scraped_posts=scraped_topics, client_inputs=client_inputs, mock=True
            )
            all_bucket_lists.append((bucket_type, bucket_topics))
        print(f"  [mock] 3 buckets: {', '.join(content_types)}")
    else:
        all_bucket_lists = []
        for bucket_type in content_types:
            display = bucket_generators.BUCKET_DISPLAY_NAMES.get(bucket_type, bucket_type)
            print(f"  Generating {display} bucket topics...")
            bucket_topics = bucket_generators.generate_bucket(
                bucket_type, config, week_of, day_dates,
                scraped_posts=scraped_topics, client_inputs=client_inputs, mock=False
            )
            all_bucket_lists.append((bucket_type, bucket_topics))
            print(f"    {len(bucket_topics)} topics generated")

    # Interleave: each day gets 1 topic from each bucket
    # Mon: bucket0[0], bucket1[0], bucket2[0]
    # Tue: bucket0[1], bucket1[1], bucket2[1]  etc.
    topics = []
    for day_idx in range(bucket_size):
        for bucket_type, bucket_topics in all_bucket_lists:
            if day_idx < len(bucket_topics):
                topics.append(bucket_topics[day_idx])

    # Renumber topic_num 1–21 sequentially
    for i, t in enumerate(topics):
        t["topic_num"] = i + 1

    # ── Airtable + R2 setup ───────────────────────────────────────────────────
    at_api_key = None
    at_base_id = None
    at_table_id = None
    use_airtable = False
    if HAS_CLOUD_MODULES and not skip_airtable and client_config.is_airtable_enabled(client_id):
        at_api_key = airtable_writer.get_api_key(client_id)
        at_base_id = config.get("airtable", {}).get("base_id", "")
        if at_api_key and at_base_id and not mock:
            print("  Setting up Airtable table...")
            try:
                at_table_id = airtable_writer.get_or_create_table(at_base_id, week_of, at_api_key)
                use_airtable = True
                print(f"  Airtable table: Week-{week_of} ({at_table_id})")
            except Exception as e:
                print(f"  Warning: Airtable setup failed: {e}")
        elif mock:
            print("  [mock] Skipping Airtable setup")
    topic_at_records = {}  # topic_num -> {"twitter": rec_id, "telegram": rec_id}

    # Print topic schedule
    print(f"\n  {'#':<4} {'Bucket':<16} {'Day':<5} {'Date':<12} {'Angle':<18} Topic")
    print(f"  {'-'*80}")
    for t in topics:
        bucket_label = bucket_generators.BUCKET_DISPLAY_NAMES.get(t.get("bucket",""), t.get("bucket",""))
        print(
            f"  {t['topic_num']:<4} {bucket_label:<16} {t['day']:<5} {t.get('date', ''):<12} "
            f"{t.get('angle', ''):<18} {t['topic'][:40]}"
        )
    print()

    # ── Phase 3: Create Workbook (opt-in via --export-excel) ──────────────────
    workbook_suffix = "mock-weekly-content" if mock else "weekly-content"
    mock_flag = " --mock" if mock else ""
    if export_excel:
        print("Phase 3: Creating workbook...")
        try:
            run_cmd(
                f"{PY} scripts/weekly_pipeline.py "
                f"--action create-workbook --week-of {week_of} --client {client_id}{mock_flag}"
            )
            print(f"  Workbook: outputs/content/{client_id}/{week_of}-{workbook_suffix}.xlsx")
        except Exception as e:
            print(f"  Error creating workbook: {e}")
            raise
    else:
        print("Phase 3: Skipped (Airtable-primary mode; use --export-excel to generate Excel)")

    # ── Phase 4: Content Generation Loop ─────────────────────────────────────
    print(f"\nPhase 4: Generating 42 content items (21 topics x 2 platforms)...")
    content_items = []

    for topic_data in topics:
        topic_num = topic_data["topic_num"]
        day = topic_data["day"]
        date = topic_data.get("date", day_dates.get(day, week_of))
        topic = topic_data["topic"]
        angle = topic_data.get("angle", "Education")

        # Position within the day (1, 2, or 3) — topics 1-2/day get thread format
        day_position = ((topic_num - 1) % 3) + 1
        twitter_fmt = "thread" if day_position <= 2 else "single"

        slug = topic_slug(topic)
        style_key, style_desc = get_style_preset(config, angle)

        for platform in ["Twitter", "Telegram"]:
            item_num = (topic_num - 1) * 2 + (1 if platform == "Twitter" else 2)
            fmt = twitter_fmt if platform == "Twitter" else "long-form"

            print(f"  Generating {item_num}/42: {day} #{topic_num} {platform} ({fmt})...")

            en_path = make_image_path(client_id, week_of, day, slug, platform)
            ru_path = make_image_path(client_id, week_of, day, slug, platform, ru=True)

            if mock:
                c = dict(MOCK_CONTENT)
                c["content"] = f"[Mock {platform} EN] {topic[:60]}"
                c["content_ru"] = f"[Мок {platform} RU] {topic[:60]}"
            else:
                format_desc = fmt + (" (5 tweets separated by ---)" if fmt == "thread" else "")
                platform_rules = get_platform_rules(config, platform, fmt)

                # Build source context from scraped post (trending bucket only)
                source_post = topic_data.get("source_post")
                if source_post:
                    sp_platform = source_post.get("platform", "social media").upper()
                    sp_sub = f" r/{source_post['subreddit']}" if source_post.get("subreddit") else ""
                    sp_eng = source_post.get("engagement", 0)
                    sp_text = source_post.get("text", "")[:400]
                    sp_url = source_post.get("url", "")
                    source_context = (
                        f"Source post from {sp_platform}{sp_sub} ({sp_eng} engagement):\n"
                        f"\"{sp_text}\"\n"
                        f"URL: {sp_url}\n\n"
                        f"Use this real community discussion as context and inspiration. "
                        f"Reference the sentiment, concerns, or angle discussed — but write original content for {display_name}.\n\n"
                    )
                else:
                    source_context = ""

                prompt = CONTENT_GEN_PROMPT.format(
                    display_name=display_name,
                    tone=tone,
                    voice=voice,
                    tagline=config.get("tagline", ""),
                    pillars=pillars,
                    ctas=ctas,
                    hashtags=hashtags,
                    topic_num=topic_num,
                    topic=topic,
                    angle=angle,
                    day=day,
                    date=date,
                    platform=platform,
                    format_desc=format_desc,
                    source_context=source_context,
                    platform_rules=platform_rules,
                    mascot=mascot,
                    bg_style=bg_style,
                    style_preset=style_desc,
                )

                try:
                    response = call_gemini(prompt, client_id=client_id)
                    c = extract_json(response)
                except Exception as e:
                    print(f"    Warning: content generation failed for item {item_num}: {e}")
                    errors.append(f"Item {item_num} content: {e}")
                    c = dict(MOCK_CONTENT)
                    c["content"] = f"[Fallback EN] {topic}"
                    c["content_ru"] = f"[Запасной RU] {topic}"

            content_json = {
                "date": date,
                "bucket": topic_data.get("bucket", "trending"),
                "day": day,
                "topic": topic,
                "platform": platform,
                "format": fmt,
                "content": c.get("content", ""),
                "image_prompt": c.get("image_prompt", f"{display_name} branded image: {topic}"),
                "image_path": en_path,
                "image_url_en": None,  # filled in Phase 5 after R2 upload
                "hashtags": c.get("hashtags", []),
                "content_ru": c.get("content_ru", ""),
                "image_prompt_ru": c.get("image_prompt_ru", f"{display_name} изображение: {topic}"),
                "image_path_ru": ru_path,
                "image_url_ru": None,  # filled in Phase 5 after R2 upload
                "hashtags_ru": c.get("hashtags_ru", []),
                "status": "Draft",
            }

            # Validate Russian content is actually Cyrillic
            ru_text = content_json.get("content_ru", "")
            if ru_text and not ru_text.startswith("[") and not is_cyrillic(ru_text):
                logger.warning(f"    RU content for item {item_num} appears non-Cyrillic, retrying...")
                try:
                    retry_prompt = (
                        f"Translate the following to Russian using Cyrillic script. "
                        f"You MUST respond entirely in Russian:\n\n{content_json['content']}"
                    )
                    retry_ru = call_gemini(retry_prompt, client_id=client_id)
                    if is_cyrillic(retry_ru):
                        content_json["content_ru"] = retry_ru
                        logger.info(f"    RU retry succeeded for item {item_num}")
                    else:
                        content_json["content_ru"] = f"[RU-WARN] {ru_text}"
                        logger.warning(f"    RU retry still non-Cyrillic for item {item_num}")
                except Exception as e:
                    logger.warning(f"    RU retry failed for item {item_num}: {e}")

            # Write to Airtable inline (both Twitter and Telegram rows)
            if use_airtable and at_table_id:
                try:
                    at_rec_id = airtable_writer.write_record(
                        at_base_id, at_table_id, content_json, week_of, client_id, at_api_key
                    )
                    if topic_num not in topic_at_records:
                        topic_at_records[topic_num] = {}
                    topic_at_records[topic_num][platform.lower()] = at_rec_id
                    print(f"    AT: {at_rec_id}")
                    time.sleep(0.2)  # Airtable rate limit
                except Exception as e:
                    print(f"    Warning: Airtable write failed: {e}")
                    errors.append(f"AT write item {item_num}: {e}")

            # Write to Excel (only if --export-excel)
            if export_excel:
                tmp_file = f"/tmp/pipeline_content_{item_num}.json"
                with open(tmp_file, "w", encoding="utf-8") as f:
                    json.dump(content_json, f, ensure_ascii=False, indent=2)
                try:
                    run_cmd(
                        f"{PY} scripts/weekly_pipeline.py --action save-content "
                        f"--week-of {week_of} --content-file {tmp_file} --client {client_id}{mock_flag}"
                    )
                except Exception as e:
                    print(f"    Warning: failed to save item {item_num} to Excel: {e}")
                    errors.append(f"Excel save item {item_num}: {e}")

            content_items.append(content_json)

    print(f"\n  Content: {len(content_items)} items generated, {len(errors)} errors so far")

    # ── Phase 5: Image Generation Loop ───────────────────────────────────────
    if not skip_images:
        num_workers = parallel_workers
        logger.info(f"Phase 5: Generating images ({num_workers} parallel workers)...")
        phase5_start = time.time()

        def _gen_images_for_topic(topic_data):
            """Generate EN + RU images for one topic. Returns list of error strings."""
            topic_errors = []
            topic_num = topic_data["topic_num"]
            day = topic_data["day"]
            topic = topic_data["topic"]
            angle = topic_data.get("angle", "Education")

            twitter_item = next(
                (c for c in content_items
                 if c["topic"] == topic and c["platform"] == "Twitter"),
                None,
            )
            if not twitter_item:
                logger.warning(f"No Twitter item for topic {topic_num}, skipping images")
                return topic_errors

            slug = topic_slug(topic)
            style_key, _ = get_style_preset(config, angle)
            en_img = twitter_item["image_path"]
            ru_img = twitter_item["image_path_ru"]

            img_dir = (PROJECT_ROOT / en_img).parent
            img_dir.mkdir(parents=True, exist_ok=True)

            logger.info(f"  Images {topic_num}/21 [{day}]: {topic[:42]}...")

            if not mock:
                en_prompt = twitter_item["image_prompt"].replace('"', "'")
                ru_prompt = twitter_item["image_prompt_ru"].replace('"', "'")

                # EN image
                try:
                    run_cmd(
                        f'{PY} scripts/nano_banana.py '
                        f'--prompt "{en_prompt}" '
                        f'--output "{en_img}" '
                        f'--style {style_key} '
                        f'--no-r2 '
                        f'--client {client_id}'
                    )
                    logger.info(f"    EN: {Path(en_img).name}")
                    if use_airtable and HAS_CLOUD_MODULES and r2_uploader.is_configured():
                        try:
                            en_img_abs = str(PROJECT_ROOT / en_img)
                            if Path(en_img_abs).exists():
                                r2_key_en = r2_uploader.make_key(client_id, week_of, Path(en_img).name)
                                r2_url_en = r2_uploader.upload_file(en_img_abs, r2_key_en)
                                logger.info(f"    R2 EN: {r2_url_en}")
                                for plat_key in ["twitter", "telegram"]:
                                    rec_id = topic_at_records.get(topic_num, {}).get(plat_key)
                                    if rec_id:
                                        airtable_writer.update_image_urls(
                                            at_base_id, at_table_id, rec_id,
                                            image_url_en=r2_url_en, api_key=at_api_key
                                        )
                        except Exception as e:
                            logger.warning(f"    R2/AT EN image update failed: {e}")
                            topic_errors.append(f"R2/AT EN image topic {topic_num}: {e}")
                except Exception as e:
                    logger.warning(f"    EN image failed for topic {topic_num}: {e}")
                    topic_errors.append(f"EN image topic {topic_num}: {e}")

                # RU image
                try:
                    run_cmd(
                        f'{PY} scripts/wavespeed_img.py '
                        f'--prompt "{ru_prompt}" '
                        f'--output "{ru_img}" '
                        f'--no-r2 '
                        f'--client {client_id}'
                    )
                    logger.info(f"    RU: {Path(ru_img).name}")
                    if use_airtable and HAS_CLOUD_MODULES and r2_uploader.is_configured():
                        try:
                            ru_img_abs = str(PROJECT_ROOT / ru_img)
                            if Path(ru_img_abs).exists():
                                r2_key_ru = r2_uploader.make_key(client_id, week_of, Path(ru_img).name)
                                r2_url_ru = r2_uploader.upload_file(ru_img_abs, r2_key_ru)
                                logger.info(f"    R2 RU: {r2_url_ru}")
                                for plat_key in ["twitter", "telegram"]:
                                    rec_id = topic_at_records.get(topic_num, {}).get(plat_key)
                                    if rec_id:
                                        airtable_writer.update_image_urls(
                                            at_base_id, at_table_id, rec_id,
                                            image_url_ru=r2_url_ru, api_key=at_api_key
                                        )
                        except Exception as e:
                            logger.warning(f"    R2/AT RU image update failed: {e}")
                            topic_errors.append(f"R2/AT RU image topic {topic_num}: {e}")
                except Exception as e:
                    logger.warning(f"    RU image failed for topic {topic_num}: {e}")
                    topic_errors.append(f"RU image topic {topic_num}: {e}")
            else:
                logger.info(f"    [mock] Skipping image API calls")

            return topic_errors

        with ThreadPoolExecutor(max_workers=num_workers) as pool:
            futures = {
                pool.submit(_gen_images_for_topic, td): td["topic_num"]
                for td in topics
            }
            for future in as_completed(futures):
                topic_num = futures[future]
                try:
                    topic_errors = future.result()
                    errors.extend(topic_errors)
                except Exception as e:
                    logger.error(f"Topic {topic_num} images crashed: {e}")
                    errors.append(f"Topic {topic_num} images failed: {e}")

        phase5_elapsed = time.time() - phase5_start
        logger.info(f"  Phase 5 complete in {phase5_elapsed:.0f}s")
    else:
        print("\nPhase 5: Skipping image generation (--skip-images)")

    # ── Phase 6: Finalize Workbook (only if --export-excel) ───────────────────
    if export_excel:
        print("\nPhase 6: Finalizing workbook...")
        try:
            run_cmd(
                f"{PY} scripts/weekly_pipeline.py "
                f"--action finalize --week-of {week_of} --client {client_id}{mock_flag}"
            )
            print("  Workbook finalized")
        except Exception as e:
            print(f"  Warning: finalize failed: {e}")
            errors.append(f"Finalize: {e}")
    else:
        print("\nPhase 6: Skipped (use --export-excel to generate Excel workbook)")

    # Phase 6.5 removed: Airtable writes now happen inline in Phase 4
    if use_airtable:
        print(f"\nPhase 6.5: Airtable inline writes complete ({len(topic_at_records)} topics written)")
    elif skip_airtable:
        print("\nPhase 6.5: Skipped (--skip-airtable)")
    else:
        print("\nPhase 6.5: Skipped (Airtable not enabled for this client)")

    # ── Phase 7: Build Static Site ────────────────────────────────────────────
    if not skip_deploy:
        print("\nPhase 7: Building static site...")
        if mock:
            print("  [mock] Skipping static site build")
        else:
            try:
                run_cmd(
                    f"{PY} scripts/build_static.py "
                    f"--output dist --include-admin --client {client_id}"
                )
                print("  Static site built: dist/")
            except Exception as e:
                print(f"  Warning: static site build failed: {e}")
                errors.append(f"Static build: {e}")
    else:
        print("\nPhase 7: Skipping static site build (--skip-deploy)")

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  {display_name} Pipeline Complete")
    print(f"  Week: {week_of}")
    print(f"  Content: {len(content_items)}/42 items generated")
    print(f"  Errors: {len(errors)}")
    if errors:
        print(f"\n  Error log (first 10):")
        for err in errors[:10]:
            print(f"    - {err}")
    if export_excel:
        print(f"\n  Excel:  outputs/content/{client_id}/{week_of}-{workbook_suffix}.xlsx")
    if use_airtable:
        print(f"  Airtable: Week-{week_of} table updated ({at_base_id})")
    print(f"  Images: outputs/content/{client_id}/images/{week_of}-weekly/")
    print(f"{'='*60}\n")

    return len(errors) == 0


REGEN_TRANSLATE_PROMPT = """\
You are a professional translator for {display_name}, a Web3 fintech platform.

Translate the following English social media content to Russian.
Preserve formatting exactly: tweet separators (---), line breaks, and emoji.
Keep hashtags as-is (do not translate #hashtag text).
Keep the same tone: {tone}

CRITICAL RULES:
- NEVER use em-dashes, en-dashes, or double-hyphens as punctuation
- Preserve tweet --- separators unchanged
- Translate only the text, not the structure

English content:
{content}

Return ONLY the Russian translation, nothing else."""


def run_regen_item(client_id, week_of, topic_index, regen_type, mock=False):
    """Regenerate a single topic item (image or content) and update the workbook."""
    from openpyxl import load_workbook as opxl_load

    config = client_config.load_config(client_id)
    display_name = config.get("display_name", client_id.capitalize())
    tone = config.get("tone", "educational and transparent")

    xlsx_path = (
        PROJECT_ROOT / f"outputs/content/{client_id}/{week_of}-weekly-content.xlsx"
    )
    if not xlsx_path.exists():
        print(f"  Error: workbook not found: {xlsx_path}")
        return False

    print(f"  Regen type: {regen_type}, topic index: {topic_index}, workbook: {xlsx_path.name}")

    # ── Read workbook to find rows for this topic ─────────────────────────────
    wb = opxl_load(xlsx_path)
    ws = wb.active
    header_row = None
    data_rows = []
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Date" or row[0] == "date":
            header_row = row
            continue
        if any(row):
            data_rows.append(row)

    # Build ordered list of unique topics (by first appearance)
    topic_order = []
    seen = set()
    for row in data_rows:
        t = row[2]  # Column C = Topic
        if t and t not in seen:
            seen.add(t)
            topic_order.append(t)

    if topic_index >= len(topic_order):
        print(f"  Error: topic_index {topic_index} out of range (found {len(topic_order)} topics)")
        return False

    target_topic = topic_order[topic_index]
    topic_rows = [r for r in data_rows if r[2] == target_topic]
    twitter_row = next((r for r in topic_rows if str(r[3]).lower() == "twitter"), topic_rows[0])

    print(f"  Topic: {target_topic}")

    # ── content: regenerate EN content via weekly_pipeline ────────────────────
    if regen_type == "content":
        sys.path.insert(0, str(PROJECT_ROOT / "scripts"))
        from weekly_pipeline import regenerate_topic_content
        try:
            result = regenerate_topic_content(
                str(xlsx_path), topic_index, client_id=client_id, mock=mock
            )
            print(f"  Content regenerated: twitter={result.get('twitter','')[:60]}...")
            return True
        except Exception as e:
            print(f"  Error regenerating content: {e}")
            return False

    # ── content_ru: re-translate EN content to Russian ────────────────────────
    elif regen_type == "content_ru":
        # Col F (index 5) = Content EN, Col J (index 9) = Content_RU
        en_content = twitter_row[5] if len(twitter_row) > 5 else ""
        if not en_content:
            print("  No EN content found to translate")
            return False

        if mock:
            print("  [mock] Skipping Gemini translation")
            return True

        prompt = REGEN_TRANSLATE_PROMPT.format(
            display_name=display_name, tone=tone, content=en_content
        )
        try:
            ru_text = call_gemini(prompt, client_id=client_id).strip()
        except Exception as e:
            print(f"  Error calling Gemini for translation: {e}")
            return False

        # Write new RU content to all rows for this topic (col J, index 10 in 1-based)
        wb2 = opxl_load(xlsx_path)
        ws2 = wb2.active
        rows_updated = 0
        in_data = False
        for row in ws2.iter_rows():
            vals = [c.value for c in row]
            if vals[0] == "Date" or vals[0] == "date":
                in_data = True
                continue
            if in_data and vals[2] == target_topic:
                row[9].value = ru_text  # Column J
                rows_updated += 1
        wb2.save(xlsx_path)
        print(f"  RU content updated in {rows_updated} rows")
        return rows_updated > 0

    # ── image_en: regenerate EN image via nano_banana.py ─────────────────────
    elif regen_type == "image_en":
        # Col G (index 6) = Image Prompt, Col H (index 7) = Image Path
        img_prompt = twitter_row[6] if len(twitter_row) > 6 else ""
        img_path = twitter_row[7] if len(twitter_row) > 7 else ""
        if not img_prompt or not img_path:
            print("  No EN image prompt/path found in workbook")
            return False

        img_dir = (PROJECT_ROOT / img_path).parent
        img_dir.mkdir(parents=True, exist_ok=True)

        style_key = "tech"  # default; could be derived from angle if needed
        safe_prompt = img_prompt.replace('"', "'")

        if mock:
            print(f"  [mock] Would generate EN image: {img_path}")
            return True

        try:
            run_cmd(
                f'{PY} scripts/nano_banana.py '
                f'--prompt "{safe_prompt}" '
                f'--output "{img_path}" '
                f'--style {style_key} '
                f'--client {client_id}'
            )
            print(f"  EN image regenerated: {Path(img_path).name}")
            return True
        except Exception as e:
            print(f"  Error generating EN image: {e}")
            return False

    # ── image_ru: regenerate RU image via wavespeed_img.py ───────────────────
    elif regen_type == "image_ru":
        # Col K (index 10) = Image_Prompt_RU, Col L (index 11) = Image_Path_RU
        ru_prompt = twitter_row[10] if len(twitter_row) > 10 else ""
        ru_path = twitter_row[11] if len(twitter_row) > 11 else ""
        if not ru_prompt or not ru_path:
            print("  No RU image prompt/path found in workbook")
            return False

        img_dir = (PROJECT_ROOT / ru_path).parent
        img_dir.mkdir(parents=True, exist_ok=True)

        safe_prompt = ru_prompt.replace('"', "'")

        if mock:
            print(f"  [mock] Would generate RU image: {ru_path}")
            return True

        try:
            run_cmd(
                f'{PY} scripts/wavespeed_img.py '
                f'--prompt "{safe_prompt}" '
                f'--output "{ru_path}" '
                f'--client {client_id}'
            )
            print(f"  RU image regenerated: {Path(ru_path).name}")
            return True
        except Exception as e:
            print(f"  Error generating RU image: {e}")
            return False

    else:
        print(f"  Unknown regen_type: {regen_type}")
        return False


def cleanup_old_images(images_dir, keep_weeks):
    """Delete image dirs older than keep_weeks weeks.

    Only removes dirs matching the pattern YYYY-MM-DD-weekly.
    Dirs that don't match the pattern are left untouched.
    """
    images_path = PROJECT_ROOT / images_dir if not Path(images_dir).is_absolute() else Path(images_dir)
    if not images_path.exists():
        return

    cutoff = datetime.now() - timedelta(weeks=keep_weeks)
    pattern = re.compile(r"^(\d{4}-\d{2}-\d{2})-weekly$")

    deleted = 0
    kept = 0
    for d in sorted(images_path.iterdir()):
        if not d.is_dir():
            continue
        m = pattern.match(d.name)
        if not m:
            continue  # skip dirs with unexpected names
        try:
            dir_date = datetime.strptime(m.group(1), "%Y-%m-%d")
        except ValueError:
            continue
        if dir_date < cutoff:
            import shutil as _shutil
            _shutil.rmtree(str(d))
            print(f"  Removed old image dir: {d.name}")
            deleted += 1
        else:
            kept += 1

    print(f"  Image cleanup: removed {deleted} old dir(s), kept {kept}")


def main():
    parser = argparse.ArgumentParser(
        description="Standalone end-to-end pipeline runner (Gemini-powered)"
    )
    parser.add_argument(
        "--client", default=None,
        help="Client ID (default: reads .active-client or bobe)",
    )
    parser.add_argument(
        "--week-of", default=None,
        help="Week start date YYYY-MM-DD (default: Monday of current week)",
    )
    parser.add_argument(
        "--mock", action="store_true",
        help="Dry run: skip all API calls, use mock content",
    )
    parser.add_argument(
        "--skip-images", action="store_true",
        help="Skip image generation phases",
    )
    parser.add_argument(
        "--skip-airtable", action="store_true",
        help="Skip Airtable sync",
    )
    parser.add_argument(
        "--skip-deploy", action="store_true",
        help="Skip static site build (GH Actions handles deploy as a separate step)",
    )
    parser.add_argument(
        "--export-excel", action="store_true",
        help="Also generate local Excel workbook (in addition to Airtable)",
    )
    parser.add_argument(
        "--regen-topic", type=int, default=None,
        help="Regenerate a single topic item (0-based index). Skips full pipeline.",
    )
    parser.add_argument(
        "--regen-type",
        choices=["image_en", "image_ru", "content", "content_ru"],
        default="image_en",
        help="What to regenerate: image_en, image_ru, content, content_ru",
    )
    parser.add_argument(
        "--mode", default="full",
        choices=["full", "announcement"],
        help="Pipeline mode: 'full' runs complete pipeline, 'announcement' generates only the announcement bucket for a week",
    )
    parser.add_argument(
        "--announcement-text", default=None,
        help="Announcement text for --mode announcement (used by generate-announcement.yml workflow)",
    )
    parser.add_argument(
        "--count", type=int, default=7,
        help="Number of announcement topics to generate (default 7, use 1 for testing)",
    )
    parser.add_argument(
        "--phase", default="all",
        choices=["all", "content", "images", "translation"],
        help="Which phase to run in announcement mode: all, content, images, or translation",
    )
    parser.add_argument(
        "--purge-older-than", type=int, default=12,
        help="Delete image dirs older than N weeks from outputs/content/{client}/images/ (0 = skip, default 12)",
    )
    parser.add_argument(
        "--parallel-workers", type=int, default=4,
        help="Number of parallel workers for image generation (default 4, set 1 for sequential)",
    )
    args = parser.parse_args()

    client_id = args.client or client_config.get_active_client()
    week_of = get_monday(getattr(args, "week_of", None))

    # Announcement-only mode: save input text and regenerate announcement bucket
    if args.mode == "announcement" and args.announcement_text:
        ann_count = min(args.count, 7)
        phase = args.phase
        print(f"\nAnnouncement mode: {client_id}, week {week_of}, count={ann_count}, phase={phase}")
        cfg = client_config.load_config(client_id)
        out_dir = client_config.get_output_dir(client_id)
        out_dir.mkdir(parents=True, exist_ok=True)

        # Save input text
        inputs_file = out_dir / f"{week_of}-bucket-inputs.json"
        inputs = {}
        if inputs_file.exists():
            try:
                inputs = json.loads(inputs_file.read_text(encoding="utf-8"))
            except Exception:
                pass
        inputs["announcements"] = {"text": args.announcement_text, "submitted_at": datetime.now().isoformat()}
        inputs_file.write_text(json.dumps(inputs, indent=2, ensure_ascii=False))
        print(f"  Input saved to {inputs_file}")

        day_dates_ann = {
            DAYS[i]: (datetime.strptime(week_of, "%Y-%m-%d") + timedelta(days=i)).strftime("%Y-%m-%d")
            for i in range(7)
        }

        # Phase: content or all — generate topic angles
        if phase in ("all", "content"):
            ann_topics = bucket_generators.generate_announcement_placeholders(
                cfg, week_of, day_dates_ann, args.announcement_text, mock=args.mock, count=ann_count
            )
            print(f"  Generated {len(ann_topics)} announcement topics")

            # Save topics to temp file so other phases can pick them up
            topics_file = out_dir / f"{week_of}-ann-topics.json"
            topics_file.write_text(json.dumps(ann_topics, indent=2, ensure_ascii=False))
        else:
            # Load existing topics for images/translation phases
            topics_file = out_dir / f"{week_of}-ann-topics.json"
            if topics_file.exists():
                ann_topics = json.loads(topics_file.read_text(encoding="utf-8"))[:ann_count]
            else:
                print(f"  Error: no topics file at {topics_file}. Run content phase first.")
                sys.exit(1)

        # Generate full content + images for each announcement topic
        # (Reuse the same per-topic content generation as Phase 4)
        xlsx_suffix = "mock-weekly-content" if args.mock else "weekly-content"
        xlsx_path = out_dir / f"{week_of}-{xlsx_suffix}.xlsx"

        from openpyxl import load_workbook as opxl_load

        # Update workbook with topics (if workbook exists and phase includes content)
        if xlsx_path.exists() and phase in ("all", "content"):
            wb = opxl_load(str(xlsx_path))
            ws = wb["Content"]
            for row in ws.iter_rows(min_row=2):
                bucket_cell = row[1]  # Column B = Bucket
                if str(bucket_cell.value or "").lower() == "announcements":
                    day_val = str(row[2].value or "")  # Column C = Day
                    ann_topic = next((t for t in ann_topics if t["day"] == day_val), None)
                    if ann_topic:
                        row[3].value = ann_topic["topic"]  # Column D = Topic
                        row[14].value = "Draft"            # Column O = Status
            wb.save(str(xlsx_path))
            print(f"  Workbook updated with announcement topics")

        # Config for content generation
        tone = cfg.get("content", {}).get("tone", "educational")
        voice = cfg.get("content", {}).get("voice", "")
        display_name = cfg.get("display_name", client_id)
        pillars = "; ".join(cfg.get("content", {}).get("messaging_pillars", []))
        ctas = "; ".join(cfg.get("content", {}).get("cta_examples", []))
        hashtags_cfg = ", ".join(cfg.get("content", {}).get("hashtags", []))
        mascot = cfg.get("brand", {}).get("mascot_description", "brand mascot")
        bg_style = cfg.get("brand", {}).get("background_style", "dark background")

        for topic_data in ann_topics:
            for platform in ["Twitter", "Telegram"]:
                day_pos = topic_data["topic_num"]
                twitter_fmt = "thread" if day_pos <= 2 else "single"
                fmt = twitter_fmt if platform == "Twitter" else "long-form"
                slug = topic_slug(topic_data["topic"])
                style_key, style_desc = get_style_preset(cfg, topic_data.get("angle", "Product"))

                en_path = make_image_path(client_id, week_of, topic_data["day"], slug, platform)
                ru_path = make_image_path(client_id, week_of, topic_data["day"], slug, platform, ru=True)

                # ── Content phase: generate EN content + image prompts ──
                if phase in ("all", "content"):
                    print(f"  Generating content: {topic_data['day']} {platform} [{topic_data['topic'][:40]}]...")

                    if not args.mock:
                        format_desc = fmt + (" (5 tweets separated by ---)" if fmt == "thread" else "")
                        platform_rules = get_platform_rules(cfg, platform, fmt)
                        prompt = CONTENT_GEN_PROMPT.format(
                            display_name=display_name, tone=tone, voice=voice,
                            tagline=cfg.get("tagline", ""), pillars=pillars, ctas=ctas,
                            hashtags=hashtags_cfg,
                            topic_num=topic_data["topic_num"], topic=topic_data["topic"],
                            angle=topic_data.get("angle", "Announcement"),
                            day=topic_data["day"], date=topic_data["date"],
                            platform=platform, format_desc=format_desc,
                            source_context="",
                            platform_rules=platform_rules, mascot=mascot, bg_style=bg_style,
                            style_preset=style_desc,
                        )
                        try:
                            response = call_gemini(prompt, client_id=client_id)
                            c = extract_json(response)
                        except Exception as e:
                            print(f"    Warning: content gen failed: {e}")
                            c = dict(MOCK_CONTENT)
                    else:
                        c = dict(MOCK_CONTENT)
                        c["content"] = f"[Mock announcement {platform} EN] {topic_data['topic'][:60]}"

                    # Save per-item JSON for other phases to read
                    content_json = {
                        "date": topic_data["date"],
                        "bucket": "announcements",
                        "day": topic_data["day"],
                        "topic": topic_data["topic"],
                        "platform": platform,
                        "format": fmt,
                        "content": c.get("content", ""),
                        "image_prompt": c.get("image_prompt", f"{display_name}: {topic_data['topic']}"),
                        "image_path": en_path,
                        "hashtags": c.get("hashtags", []),
                        "content_ru": c.get("content_ru", ""),
                        "image_prompt_ru": c.get("image_prompt_ru", ""),
                        "image_path_ru": ru_path,
                        "hashtags_ru": c.get("hashtags_ru", []),
                        "status": "Draft",
                    }

                    tmp_file = out_dir / f"ann_content_{topic_data['day']}_{platform.lower()}.json"
                    tmp_file.write_text(json.dumps(content_json, indent=2, ensure_ascii=False))

                    # Update workbook row
                    if xlsx_path.exists():
                        wb2 = opxl_load(str(xlsx_path))
                        ws2 = wb2["Content"]
                        for row in ws2.iter_rows(min_row=2):
                            if (str(row[1].value or "").lower() == "announcements" and
                                    str(row[2].value or "") == topic_data["day"] and
                                    str(row[4].value or "").lower() == platform.lower()):
                                row[3].value = topic_data["topic"]
                                row[6].value = c.get("content", "")
                                row[7].value = c.get("image_prompt", "")
                                row[8].value = en_path
                                row[9].value = ", ".join(c.get("hashtags", []))
                                row[10].value = c.get("content_ru", "")
                                row[11].value = c.get("image_prompt_ru", "")
                                row[12].value = ru_path
                                row[13].value = ", ".join(c.get("hashtags_ru", []))
                                row[14].value = "Draft"
                                break
                        wb2.save(str(xlsx_path))

                # ── Translation phase: regenerate only RU content ──
                if phase == "translation":
                    tmp_file = out_dir / f"ann_content_{topic_data['day']}_{platform.lower()}.json"
                    if not tmp_file.exists():
                        print(f"  Skipping translation {topic_data['day']} {platform}: no content file. Run content phase first.")
                        continue
                    c = json.loads(tmp_file.read_text(encoding="utf-8"))
                    en_content = c.get("content", "")
                    en_hashtags = c.get("hashtags", [])
                    if not en_content:
                        print(f"  Skipping translation {topic_data['day']} {platform}: no EN content.")
                        continue

                    print(f"  Translating: {topic_data['day']} {platform} [{topic_data.get('topic', '')[:40]}]...")
                    if not args.mock:
                        translate_prompt = (
                            f"Translate this social media post to Russian. Keep the same tone, formatting, and structure.\n"
                            f"Also translate these hashtags to Russian equivalents: {', '.join(en_hashtags)}\n"
                            f"Also create a Russian image prompt based on this English one: {c.get('image_prompt', '')}\n\n"
                            f"English content:\n{en_content}\n\n"
                            f'Return ONLY JSON: {{"content_ru": "...", "hashtags_ru": ["..."], "image_prompt_ru": "..."}}'
                        )
                        try:
                            resp = call_gemini(translate_prompt, client_id=client_id)
                            tr = extract_json(resp)
                        except Exception as e:
                            print(f"    Warning: translation failed: {e}")
                            tr = {"content_ru": "", "hashtags_ru": [], "image_prompt_ru": ""}
                    else:
                        tr = {
                            "content_ru": f"[Mock RU] {en_content[:60]}",
                            "hashtags_ru": ["#тест"],
                            "image_prompt_ru": f"[Mock RU prompt] {c.get('image_prompt', '')[:40]}",
                        }

                    # Update the saved content JSON
                    c["content_ru"] = tr.get("content_ru", "")
                    c["hashtags_ru"] = tr.get("hashtags_ru", [])
                    c["image_prompt_ru"] = tr.get("image_prompt_ru", "")
                    tmp_file.write_text(json.dumps(c, indent=2, ensure_ascii=False))

                    # Update workbook row
                    if xlsx_path.exists():
                        wb2 = opxl_load(str(xlsx_path))
                        ws2 = wb2["Content"]
                        for row in ws2.iter_rows(min_row=2):
                            if (str(row[1].value or "").lower() == "announcements" and
                                    str(row[2].value or "") == topic_data["day"] and
                                    str(row[4].value or "").lower() == platform.lower()):
                                row[10].value = tr.get("content_ru", "")
                                row[11].value = tr.get("image_prompt_ru", "")
                                row[13].value = ", ".join(tr.get("hashtags_ru", []))
                                break
                        wb2.save(str(xlsx_path))

                # ── Images phase: generate EN + RU images ──
                if phase in ("all", "images"):
                    tmp_file = out_dir / f"ann_content_{topic_data['day']}_{platform.lower()}.json"
                    if tmp_file.exists():
                        c = json.loads(tmp_file.read_text(encoding="utf-8"))
                    else:
                        print(f"  Skipping images {topic_data['day']} {platform}: no content file. Run content phase first.")
                        continue

                    if not args.mock and not args.skip_images:
                        img_dir = (PROJECT_ROOT / en_path).parent
                        img_dir.mkdir(parents=True, exist_ok=True)
                        safe_en = c.get("image_prompt", "").replace('"', "'")
                        safe_ru = c.get("image_prompt_ru", "").replace('"', "'")
                        print(f"  Generating images: {topic_data['day']} {platform}...")
                        try:
                            run_cmd(f'{PY} scripts/nano_banana.py --prompt "{safe_en}" --output "{en_path}" --style {style_key} --client {client_id}')
                        except Exception as e:
                            print(f"    Warning: EN image failed: {e}")
                        if safe_ru:
                            try:
                                run_cmd(f'{PY} scripts/wavespeed_img.py --prompt "{safe_ru}" --output "{ru_path}" --client {client_id}')
                            except Exception as e:
                                print(f"    Warning: RU image failed: {e}")
                    elif args.mock:
                        print(f"  [Mock] Images: {topic_data['day']} {platform}")

        print(f"\nAnnouncement mode complete. Phase: {phase}, count: {ann_count}, week: {week_of}")
        sys.exit(0)

    if args.regen_topic is not None:
        success = run_regen_item(
            client_id=client_id,
            week_of=week_of,
            topic_index=args.regen_topic,
            regen_type=args.regen_type,
            mock=args.mock,
        )
        sys.exit(0 if success else 1)

    success = run_pipeline(
        client_id=client_id,
        week_of=week_of,
        mock=args.mock,
        skip_images=args.skip_images,
        skip_airtable=args.skip_airtable,
        skip_deploy=args.skip_deploy,
        export_excel=args.export_excel,
        parallel_workers=args.parallel_workers,
    )

    if success and args.mode == "full" and args.purge_older_than > 0:
        output_dir = client_config.get_output_dir(client_id)
        cleanup_old_images(output_dir / "images", args.purge_older_than)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
