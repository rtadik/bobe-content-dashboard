#!/usr/bin/env python3
"""
BoBe Content Pipeline - Excel Manager
Creates and manages Excel spreadsheets for content pipeline tracking.

Usage:
  python scripts/excel_manager.py --action create --date 2026-02-18
  python scripts/excel_manager.py --action add-topics --file outputs/content/2026-02-18-content.xlsx --topics topics.json
  python scripts/excel_manager.py --action add-content --file outputs/content/2026-02-18-content.xlsx --content content.json
"""

import os
import json
import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

__all__ = [
    "create_daily_workbook", "add_topics_to_sheet1", "add_content_to_sheet2",
    "save_workbook", "style_header_cell", "style_data_cell",
    "COLOR_DARK_BG", "COLOR_BLUE", "COLOR_GREEN", "COLOR_WHITE", "COLOR_LIGHT_ROW",
]

OUTPUT_DIR = Path(__file__).parent.parent / "outputs" / "content"

# Colors matching BoBe brand
COLOR_DARK_BG = "111B32"
COLOR_BLUE = "1589DC"
COLOR_GREEN = "5BD69F"
COLOR_YELLOW = "E0C145"
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_ROW = "1A2540"


def style_header_cell(cell, bg_color: str = COLOR_BLUE, text_color: str = COLOR_WHITE):
    """Apply header styling to a cell."""
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.font = Font(bold=True, color=text_color, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_cell(cell, row_index: int):
    """Apply alternating row styling."""
    bg = COLOR_DARK_BG if row_index % 2 == 0 else COLOR_LIGHT_ROW
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=COLOR_WHITE, size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def create_daily_workbook(date: str) -> Workbook:
    """Create a new workbook with Topics and Content sheets."""
    wb = Workbook()

    # --- Sheet 1: Topics ---
    ws1 = wb.active
    ws1.title = "Topics"

    topic_headers = [
        "Date", "Platform", "Topic Summary", "Original Text",
        "Author", "URL", "Engagement Score", "Relevance Score", "Selected for Content"
    ]
    topic_col_widths = [12, 10, 35, 60, 20, 40, 16, 16, 20]

    for i, (header, width) in enumerate(zip(topic_headers, topic_col_widths), 1):
        cell = ws1.cell(row=1, column=i, value=header)
        style_header_cell(cell)
        ws1.column_dimensions[get_column_letter(i)].width = width

    ws1.row_dimensions[1].height = 35
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Content ---
    ws2 = wb.create_sheet("Content")

    content_headers = [
        "Date", "Topic", "Platform Target", "Format",
        "Content", "Image Prompt", "Image Path", "Hashtags", "Status"
    ]
    content_col_widths = [12, 30, 14, 10, 70, 50, 35, 35, 12]

    for i, (header, width) in enumerate(zip(content_headers, content_col_widths), 1):
        cell = ws2.cell(row=1, column=i, value=header)
        style_header_cell(cell)
        ws2.column_dimensions[get_column_letter(i)].width = width

    ws2.row_dimensions[1].height = 35
    ws2.freeze_panes = "A2"

    return wb


def add_topics_to_sheet1(workbook: Workbook, topics: list) -> None:
    """Add scraped topics to the Topics sheet."""
    ws = workbook["Topics"]
    next_row = ws.max_row + 1

    for i, topic in enumerate(topics):
        row_idx = next_row + i
        values = [
            topic.get("date", datetime.now().strftime("%Y-%m-%d")),
            topic.get("platform", "").upper(),
            topic.get("topic_summary", topic.get("text", "")[:100]),
            topic.get("text", ""),
            topic.get("author", ""),
            topic.get("url", ""),
            topic.get("engagement", 0),
            topic.get("relevance_score", 0),
            "No",
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_data_cell(cell, i)

        ws.row_dimensions[row_idx].height = 60


def add_content_to_sheet2(workbook: Workbook, content: list) -> None:
    """Add generated content to the Content sheet."""
    ws = workbook["Content"]
    next_row = ws.max_row + 1

    for i, item in enumerate(content):
        row_idx = next_row + i
        values = [
            item.get("date", datetime.now().strftime("%Y-%m-%d")),
            item.get("topic", ""),
            item.get("platform", ""),
            item.get("format", ""),
            item.get("content", ""),
            item.get("image_prompt", ""),
            item.get("image_path", ""),
            ", ".join(item.get("hashtags", [])),
            item.get("status", "Draft"),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_data_cell(cell, i)

        ws.row_dimensions[row_idx].height = 80


def save_workbook(workbook: Workbook, date: str, output_dir: Path = OUTPUT_DIR) -> str:
    """Save workbook to outputs/content/ with dated filename."""
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{date}-content.xlsx"
    filepath = output_dir / filename
    workbook.save(str(filepath))
    return str(filepath)


def main():
    parser = argparse.ArgumentParser(description="Manage BoBe content pipeline Excel files")
    parser.add_argument("--action", choices=["create", "add-topics", "add-content"],
                        required=True, help="Action to perform")
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"),
                        help="Date for the workbook (YYYY-MM-DD)")
    parser.add_argument("--file", default=None,
                        help="Path to existing workbook (for add-topics, add-content)")
    parser.add_argument("--topics", default=None,
                        help="Path to topics JSON file")
    parser.add_argument("--content", default=None,
                        help="Path to content JSON file")
    parser.add_argument("--mock", action="store_true",
                        help="Use mock data for testing")

    args = parser.parse_args()

    if args.action == "create":
        print(f"Creating workbook for {args.date}...")
        wb = create_daily_workbook(args.date)

        if args.mock:
            # Add sample data in mock mode
            mock_topics = [
                {
                    "platform": "twitter",
                    "text": "DCA bots have been printing quietly while everyone panics. Automation > emotion #DeFi",
                    "author": "crypto_trader_x",
                    "url": "https://twitter.com/mock/1",
                    "engagement": 603,
                    "relevance_score": 3,
                    "date": args.date,
                },
            ]
            add_topics_to_sheet1(wb, mock_topics)
            print("  Added mock topic data")

        filepath = save_workbook(wb, args.date)
        print(f"Workbook created: {filepath}")

    elif args.action == "add-topics":
        if not args.file:
            print("Error: --file required for add-topics action")
            return
        if not args.topics and not args.mock:
            print("Error: --topics or --mock required")
            return

        wb = load_workbook(args.file)
        if args.mock:
            topics = [{"platform": "twitter", "text": "Mock topic", "engagement": 100, "relevance_score": 2}]
        else:
            with open(args.topics) as f:
                topics = json.load(f)

        add_topics_to_sheet1(wb, topics)
        wb.save(args.file)
        print(f"Added {len(topics)} topics to {args.file}")

    elif args.action == "add-content":
        if not args.file:
            print("Error: --file required for add-content action")
            return
        if not args.content and not args.mock:
            print("Error: --content or --mock required")
            return

        wb = load_workbook(args.file)
        if args.mock:
            content = [{
                "topic": "Automation vs emotion",
                "platform": "Twitter",
                "format": "thread",
                "content": "Mock generated content here",
                "image_prompt": "Mock image prompt",
                "hashtags": ["#DeFi", "#TradingBot"],
                "status": "Draft",
            }]
        else:
            with open(args.content) as f:
                content = json.load(f)

        add_content_to_sheet2(wb, content)
        wb.save(args.file)
        print(f"Added {len(content)} content items to {args.file}")


if __name__ == "__main__":
    main()
