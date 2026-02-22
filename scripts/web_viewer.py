#!/usr/bin/env python3
"""
Content Dashboard — Flask web viewer
Multi-client: reads brand name and output paths from client config.
Serves a bilingual (EN/RU) visual dashboard at http://localhost:5001

Usage:
  python scripts/web_viewer.py                       # auto-loads latest weekly workbook
  python scripts/web_viewer.py week:2026-02-16       # loads specific week
  python scripts/web_viewer.py --client newclient     # view a specific client
"""

import sys
import re
import json
import time
import uuid
import threading
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

import client_config

# Resolve active client for this session
_active_client = client_config.get_active_client()
_client_config = client_config.load_config(_active_client)
_display_name = _client_config.get("display_name", _active_client)

CONTENT_DIR  = client_config.get_output_dir(_active_client)
IMAGES_DIR   = CONTENT_DIR / "images"


def _reload_active_client(client_id):
    """Update all module-level globals and persist the active client switch."""
    global _active_client, _client_config, _display_name, CONTENT_DIR, IMAGES_DIR
    client_config.set_active_client(client_id)
    _active_client = client_id
    _client_config = client_config.load_config(client_id)
    _display_name = _client_config.get("display_name", client_id)
    CONTENT_DIR = client_config.get_output_dir(client_id)
    IMAGES_DIR = CONTENT_DIR / "images"


try:
    from flask import Flask, render_template_string, send_from_directory, request, jsonify
except ImportError:
    print("Flask not installed. Run:  venv/bin/pip install flask")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run:  venv/bin/pip install openpyxl")
    sys.exit(1)

# Try to import image generator — graceful degradation if not available
try:
    from nano_banana import generate_image as gen_img, build_prompt, get_default_references
    HAS_GENERATOR = True
except Exception:
    HAS_GENERATOR = False

try:
    from weekly_pipeline import translate_text_to_russian, translate_hashtags_to_russian, update_ru_columns
    from wavespeed_img import translate_image as wavespeed_translate_image
    HAS_RU_GENERATOR = True
except Exception:
    HAS_RU_GENERATOR = False

app = Flask(__name__)

# ── Async job tracking ────────────────────────────────────────────────────────
# { job_id: {"status": "running"|"done"|"error", "filename": str, "error": str} }
_jobs: dict = {}
_jobs_lock = threading.Lock()


# ── Excel discovery ───────────────────────────────────────────────────────────

def list_available_dates():
    dates = []
    # Weekly files only: YYYY-MM-DD-weekly-content.xlsx — prefix with "week:"
    for f in sorted(CONTENT_DIR.glob("*-weekly-content.xlsx"), reverse=True):
        m = re.match(r"(\d{4}-\d{2}-\d{2})-weekly-content\.xlsx", f.name)
        if m:
            dates.append(f"week:{m.group(1)}")
    return dates


def find_excel(date=None):
    if date and date.startswith("week:"):
        week_of = date[len("week:"):]
        p = CONTENT_DIR / f"{week_of}-weekly-content.xlsx"
        return p if p.exists() else None
    # Default: most recent weekly workbook
    dates = list_available_dates()
    if not dates:
        return None
    week_of = dates[0][len("week:"):]
    return CONTENT_DIR / f"{week_of}-weekly-content.xlsx"


# ── Image resolution ──────────────────────────────────────────────────────────

def resolve_image(raw_path, topic):
    """Return path relative to IMAGES_DIR of the best matching image, or None."""
    if not raw_path:
        return None

    # Strategy 1: exact path relative to project root
    candidate = PROJECT_ROOT / raw_path
    if candidate.exists():
        try:
            return str(candidate.relative_to(IMAGES_DIR))
        except ValueError:
            return candidate.name

    # Strategy 2: exact filename match — search recursively in IMAGES_DIR
    filename = Path(raw_path).name
    for png in sorted(IMAGES_DIR.glob("**/*.png")):
        if png.name == filename:
            return str(png.relative_to(IMAGES_DIR))
    if (CONTENT_DIR / filename).exists():
        return filename

    # Strategy 3: fuzzy slug match
    stem = Path(raw_path).stem
    parts = stem.split("_")
    slug_parts = [p for p in parts[1:] if p not in ("twitter", "telegram", "v2", "v1", "ru")]
    slug = "_".join(slug_parts)
    if slug:
        for png in sorted(IMAGES_DIR.glob("**/*.png")):
            if slug in png.name:
                return str(png.relative_to(IMAGES_DIR))

    # Strategy 4: derive slug from topic text
    topic_slug = re.sub(r"[^a-z0-9]+", "_", topic.lower())[:30].strip("_")
    for word in topic_slug.split("_"):
        if len(word) > 3:
            for png in sorted(IMAGES_DIR.glob("**/*.png")):
                if word in png.name.lower():
                    return str(png.relative_to(IMAGES_DIR))

    return None


# ── Approval state management ─────────────────────────────────────────────────

def get_approval_file(date):
    return CONTENT_DIR / f"{date}-approvals.json"


def load_approvals(date):
    f = get_approval_file(date)
    if f.exists():
        try:
            return json.loads(f.read_text())
        except Exception:
            return {}
    return {}


def save_approvals(date, approvals):
    f = get_approval_file(date)
    f.write_text(json.dumps(approvals, indent=2))


# ── Data loading ──────────────────────────────────────────────────────────────

def load_content(xlsx_path):
    """
    Read the Content sheet and return list of topic dicts grouped by topic.
    Supports 10-column (old weekly) and 14-column (bilingual weekly) workbooks.
    Each dict: {topic, date, day, img_prompt, image_filename, img_prompt_ru,
                image_filename_ru, twitter, telegram, twitter_ru, telegram_ru,
                hashtag_list, hashtag_list_ru}
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Content" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["Content"]
    topics = {}
    topic_order = []

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    num_cols = len([h for h in header_row if h])
    has_ru = num_cols >= 14

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue

        if has_ru:
            (date, day, topic, platform, fmt, content, img_prompt, img_path, hashtags,
             content_ru, img_prompt_ru, img_path_ru, hashtags_ru, status) = row[:14]
        else:
            # Old 10-column weekly (graceful degradation — no RU fields)
            date, day, topic, platform, fmt, content, img_prompt, img_path, hashtags, status = row[:10]
            content_ru = img_prompt_ru = img_path_ru = hashtags_ru = ""

        if topic not in topics:
            topics[topic] = {
                "topic":            topic,
                "date":             str(date) if date else "",
                "day":              day or "",
                "img_prompt":       img_prompt or "",
                "raw_image_path":   img_path or "",
                "image_filename":   None,
                "img_prompt_ru":    img_prompt_ru or "",
                "raw_image_path_ru": img_path_ru or "",
                "image_filename_ru": None,
                "twitter":          None,
                "telegram":         None,
                "twitter_ru":       None,
                "telegram_ru":      None,
                "hashtags":         hashtags or "",
                "hashtags_ru":      hashtags_ru or "",
            }
            topic_order.append(topic)

        platform_lower = (platform or "").lower()
        if "twitter" in platform_lower:
            topics[topic]["twitter"] = content or ""
            if content_ru:
                topics[topic]["twitter_ru"] = content_ru or ""
        elif "telegram" in platform_lower:
            topics[topic]["telegram"] = content or ""
            if hashtags:
                topics[topic]["hashtags"] = hashtags
            if content_ru:
                topics[topic]["telegram_ru"] = content_ru or ""
            if hashtags_ru:
                topics[topic]["hashtags_ru"] = hashtags_ru

    wb.close()

    result = []
    for key in topic_order:
        t = topics[key]
        t["image_filename"]    = resolve_image(t["raw_image_path"],    t["topic"])
        t["image_filename_ru"] = resolve_image(t["raw_image_path_ru"], t["topic"])
        raw = t["hashtags"]
        t["hashtag_list"] = [h.strip() for h in str(raw).split(",") if h.strip()] if raw else []
        raw_ru = t["hashtags_ru"]
        t["hashtag_list_ru"] = [h.strip() for h in str(raw_ru).split(",") if h.strip()] if raw_ru else []
        result.append(t)

    return result


# ── HTML template ─────────────────────────────────────────────────────────────

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ brand_name }} Content Dashboard</title>
<link rel="icon" type="image/jpeg" href="/favicon.jpg">
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  :root {
    --bg:        #0D1526;
    --surface:   #111B32;
    --surface2:  #162038;
    --border:    rgba(21,137,220,0.15);
    --blue:      #1589DC;
    --blue-dim:  rgba(21,137,220,0.12);
    --green:     #5BD69F;
    --green-dim: rgba(91,214,159,0.12);
    --yellow:    #E0C145;
    --yellow-dim:rgba(224,193,69,0.12);
    --pink:      #FF4FDA;
    --white:     #FFFFFF;
    --muted:     #6B82A8;
    --text:      #C8D8EE;
  }

  html, body {
    background: var(--bg);
    color: var(--white);
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Inter', sans-serif;
    min-height: 100vh;
    line-height: 1.5;
  }

  /* ── Header ── */
  header {
    background: #080F1E;
    border-bottom: 1px solid var(--border);
    padding: 0 24px;
    height: 60px;
    display: flex;
    align-items: center;
    gap: 16px;
    position: sticky;
    top: 0;
    z-index: 200;
  }

  .brand {
    display: flex;
    align-items: center;
    gap: 8px;
    font-size: 1.1rem;
    font-weight: 700;
    color: var(--blue);
    letter-spacing: 0.3px;
    white-space: nowrap;
  }

  .brand-dot { color: var(--pink); }

  .header-sep {
    width: 1px;
    height: 20px;
    background: var(--border);
  }

  .header-date {
    font-size: 0.85rem;
    color: var(--muted);
    white-space: nowrap;
  }

  .header-count {
    font-size: 0.78rem;
    background: var(--blue-dim);
    border: 1px solid rgba(21,137,220,0.25);
    color: var(--blue);
    padding: 2px 9px;
    border-radius: 20px;
    white-space: nowrap;
  }

  .spacer { flex: 1; }

  .date-select {
    background: var(--surface2);
    border: 1px solid var(--border);
    color: var(--text);
    padding: 5px 10px;
    border-radius: 7px;
    font-size: 0.82rem;
    cursor: pointer;
    outline: none;
  }
  .date-select:hover { border-color: var(--blue); }

  /* ── Admin link ── */
  .admin-link {
    color: var(--muted);
    text-decoration: none;
    font-size: 0.82rem;
    font-weight: 500;
    padding: 5px 10px;
    border-radius: 7px;
    transition: color 0.15s, background 0.15s;
    white-space: nowrap;
  }
  .admin-link:hover {
    color: var(--blue);
    background: var(--blue-dim);
  }

  /* ── Language toggle ── */
  .lang-toggle {
    display: flex;
    background: var(--surface2);
    border: 1px solid var(--border);
    border-radius: 8px;
    overflow: hidden;
  }
  .lang-btn {
    background: none;
    border: none;
    color: var(--muted);
    padding: 5px 14px;
    cursor: pointer;
    font-size: 0.78rem;
    font-weight: 600;
    letter-spacing: 0.5px;
    transition: all 0.15s;
  }
  .lang-btn.active {
    background: var(--blue);
    color: #fff;
  }
  .lang-btn:hover:not(.active) { color: var(--text); }

  /* EN/RU visibility — hide RU elements by default, show in RU mode */
  .ru-only,
  div.ru-only,
  .card-img.ru-only,
  .image-actions.ru-only { display: none !important; }
  body.lang-ru .ru-only,
  body.lang-ru div.ru-only { display: block !important; }
  body.lang-ru .card-img.ru-only { display: block !important; }
  body.lang-ru .image-actions.ru-only { display: flex !important; }
  body.lang-ru .en-only,
  body.lang-ru div.en-only,
  body.lang-ru .card-img.en-only,
  body.lang-ru .image-actions.en-only { display: none !important; }

  /* ── Grid ── */
  .grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(400px, 1fr));
    gap: 24px;
    padding: 28px 24px;
    max-width: 1440px;
    margin: 0 auto;
  }

  @media (max-width: 500px) {
    .grid { grid-template-columns: 1fr; padding: 14px; gap: 16px; }
    header { padding: 0 14px; gap: 10px; }
  }

  /* ── Card ── */
  .card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 14px;
    overflow: hidden;
    display: flex;
    flex-direction: column;
    transition: border-color 0.2s, box-shadow 0.2s;
  }
  .card:hover {
    border-color: rgba(21,137,220,0.35);
    box-shadow: 0 4px 32px rgba(21,137,220,0.08);
  }
  .card.is-approved {
    border-color: rgba(91,214,159,0.4);
  }
  .card.is-approved:hover {
    border-color: rgba(91,214,159,0.6);
    box-shadow: 0 4px 32px rgba(91,214,159,0.08);
  }

  /* ── Card image ── */
  .card-img {
    position: relative;
    width: 100%;
    aspect-ratio: 16 / 9;
    overflow: hidden;
    background: var(--surface2);
    cursor: pointer;
  }
  .card-img img {
    width: 100%;
    height: 100%;
    object-fit: cover;
    display: block;
    transition: transform 0.3s ease;
  }
  .card-img:hover img { transform: scale(1.02); }
  .card-img .img-overlay {
    position: absolute;
    inset: 0;
    background: linear-gradient(to bottom, transparent 60%, rgba(13,21,38,0.7) 100%);
    pointer-events: none;
  }
  .no-img {
    width: 100%;
    height: 100%;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    gap: 6px;
    color: var(--muted);
    font-size: 0.8rem;
  }
  .no-img-icon { font-size: 2rem; opacity: 0.3; }

  /* ── Loading overlay ── */
  .img-loading-overlay {
    position: absolute;
    inset: 0;
    background: rgba(13,21,38,0.88);
    display: none;
    align-items: center;
    justify-content: center;
    flex-direction: column;
    gap: 12px;
    z-index: 10;
    pointer-events: all;
  }
  .img-loading-overlay.active { display: flex; }
  .spinner {
    width: 38px;
    height: 38px;
    border: 3px solid rgba(21,137,220,0.2);
    border-top-color: var(--blue);
    border-radius: 50%;
    animation: spin 0.8s linear infinite;
  }
  @keyframes spin { to { transform: rotate(360deg); } }
  .spinner-label {
    font-size: 0.75rem;
    color: var(--muted);
    letter-spacing: 0.3px;
  }

  /* ── Image action bar ── */
  .image-actions {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 8px 12px;
    background: #0A1221;
    border-bottom: 1px solid var(--border);
    min-height: 40px;
  }

  .img-status {
    flex: 1;
    font-size: 0.72rem;
    color: var(--muted);
    letter-spacing: 0.2px;
  }
  .img-status.approved {
    color: var(--green);
    font-weight: 500;
  }

  .action-btn {
    border: 1px solid transparent;
    padding: 4px 12px;
    border-radius: 7px;
    cursor: pointer;
    font-size: 0.75rem;
    font-weight: 500;
    transition: all 0.15s;
    white-space: nowrap;
  }
  .action-btn:disabled {
    opacity: 0.45;
    cursor: not-allowed;
  }

  .approve-btn {
    background: var(--green-dim);
    border-color: rgba(91,214,159,0.25);
    color: var(--green);
  }
  .approve-btn:hover:not(:disabled) {
    background: rgba(91,214,159,0.2);
    border-color: rgba(91,214,159,0.4);
  }
  .approve-btn.approved {
    background: var(--green);
    border-color: var(--green);
    color: #0D1526;
  }
  .approve-btn.approved:hover:not(:disabled) {
    background: #4BC48F;
  }

  .regen-btn {
    background: rgba(255,79,218,0.07);
    border-color: rgba(255,79,218,0.2);
    color: var(--pink);
  }
  .regen-btn:hover:not(:disabled) {
    background: rgba(255,79,218,0.14);
    border-color: rgba(255,79,218,0.35);
  }

  /* ── Card body ── */
  .card-body {
    padding: 16px 18px 18px;
    flex: 1;
    display: flex;
    flex-direction: column;
    gap: 14px;
  }

  .topic-title {
    font-size: 0.9rem;
    font-weight: 600;
    color: var(--white);
    line-height: 1.45;
  }

  /* ── Platform tabs ── */
  .tab-bar {
    display: flex;
    gap: 6px;
  }

  .tab {
    background: none;
    border: 1px solid var(--border);
    color: var(--muted);
    padding: 5px 14px;
    border-radius: 8px;
    cursor: pointer;
    font-size: 0.78rem;
    font-weight: 500;
    letter-spacing: 0.3px;
    transition: all 0.15s;
    display: flex;
    align-items: center;
    gap: 5px;
  }
  .tab:hover:not(.active) {
    border-color: rgba(21,137,220,0.3);
    color: var(--text);
    background: var(--blue-dim);
  }
  .tab.active {
    background: var(--blue-dim);
    border-color: rgba(21,137,220,0.4);
    color: var(--blue);
  }
  .tab .platform-icon { font-size: 0.9rem; }

  /* ── Content panels ── */
  .panel { display: none; flex-direction: column; gap: 8px; }
  .panel.active { display: flex; }

  .content-box {
    background: #080F1E;
    border: 1px solid var(--border);
    border-radius: 9px;
    padding: 13px 14px;
    font-size: 0.8rem;
    line-height: 1.65;
    color: var(--text);
    white-space: pre-wrap;
    word-wrap: break-word;
    max-height: 220px;
    overflow-y: auto;
    font-family: inherit;
    scrollbar-width: thin;
    scrollbar-color: var(--border) transparent;
  }
  .content-box::-webkit-scrollbar { width: 4px; }
  .content-box::-webkit-scrollbar-track { background: transparent; }
  .content-box::-webkit-scrollbar-thumb {
    background: var(--border);
    border-radius: 4px;
  }

  .tweet-divider {
    border: none;
    border-top: 1px dashed rgba(21,137,220,0.2);
    margin: 6px 0;
  }

  .content-actions {
    display: flex;
    align-items: center;
    justify-content: space-between;
  }

  .char-count {
    font-size: 0.72rem;
    color: var(--muted);
  }

  .copy-btn {
    background: var(--blue);
    color: #fff;
    border: none;
    padding: 5px 14px;
    border-radius: 7px;
    cursor: pointer;
    font-size: 0.77rem;
    font-weight: 500;
    transition: background 0.15s, transform 0.1s;
  }
  .copy-btn:hover { background: #1070BB; }
  .copy-btn:active { transform: scale(0.97); }
  .copy-btn.copied { background: var(--green); color: #0D1526; }

  /* ── Hashtags ── */
  .hashtags {
    display: flex;
    flex-wrap: wrap;
    gap: 5px;
    padding-top: 2px;
  }
  .tag {
    background: var(--yellow-dim);
    color: var(--yellow);
    border: 1px solid rgba(224,193,69,0.25);
    border-radius: 20px;
    padding: 2px 9px;
    font-size: 0.72rem;
    cursor: pointer;
    transition: background 0.15s;
    user-select: none;
  }
  .tag:hover { background: rgba(224,193,69,0.2); }
  .tag.tag-copied { background: var(--green-dim); color: var(--green); border-color: rgba(91,214,159,0.3); }

  /* ── Lightbox ── */
  #lightbox {
    display: none;
    position: fixed;
    inset: 0;
    background: rgba(8,15,30,0.93);
    z-index: 1000;
    align-items: center;
    justify-content: center;
    cursor: zoom-out;
    padding: 20px;
  }
  #lightbox.open { display: flex; }
  #lightbox img {
    max-width: 90vw;
    max-height: 85vh;
    border-radius: 10px;
    box-shadow: 0 8px 60px rgba(0,0,0,0.7);
    cursor: default;
  }
  #lightbox-close {
    position: absolute;
    top: 18px;
    right: 22px;
    background: none;
    border: none;
    color: var(--muted);
    font-size: 1.6rem;
    cursor: pointer;
    line-height: 1;
    padding: 4px 8px;
    border-radius: 6px;
  }
  #lightbox-close:hover { color: var(--white); background: var(--surface2); }

  /* ── Toast ── */
  #toast {
    position: fixed;
    bottom: 24px;
    left: 50%;
    transform: translateX(-50%) translateY(20px);
    background: #1A2540;
    border: 1px solid var(--border);
    color: var(--text);
    padding: 9px 18px;
    border-radius: 9px;
    font-size: 0.8rem;
    opacity: 0;
    transition: opacity 0.2s, transform 0.2s;
    pointer-events: none;
    z-index: 500;
    white-space: nowrap;
  }
  #toast.show {
    opacity: 1;
    transform: translateX(-50%) translateY(0);
  }
  #toast.error {
    border-color: rgba(255,79,218,0.3);
    color: #FF9EED;
  }

  /* ── Empty state ── */
  .empty {
    grid-column: 1 / -1;
    text-align: center;
    padding: 60px 20px;
    color: var(--muted);
  }
  .empty h2 { font-size: 1.1rem; margin-bottom: 8px; color: var(--text); }
  .empty code {
    background: var(--surface2);
    padding: 2px 7px;
    border-radius: 5px;
    font-size: 0.85rem;
    color: var(--blue);
  }

  /* ── Generate RU button ── */
  .generate-ru-bar {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 10px 14px;
    background: rgba(21,137,220,0.05);
    border: 1px dashed rgba(21,137,220,0.2);
    border-radius: 9px;
    margin-bottom: 4px;
  }
  .generate-ru-bar .ru-gen-label {
    flex: 1;
    font-size: 0.78rem;
    color: var(--muted);
  }
  .generate-ru-bar .ru-gen-label.done {
    color: var(--green);
  }
  .generate-ru-btn {
    background: var(--blue-dim);
    border: 1px solid rgba(21,137,220,0.3);
    color: var(--blue);
    padding: 5px 14px;
    border-radius: 7px;
    cursor: pointer;
    font-size: 0.78rem;
    font-weight: 500;
    transition: all 0.15s;
    white-space: nowrap;
  }
  .generate-ru-btn:hover:not(:disabled) {
    background: rgba(21,137,220,0.2);
    border-color: rgba(21,137,220,0.5);
  }
  .generate-ru-btn:disabled {
    opacity: 0.4;
    cursor: not-allowed;
  }
  .generate-ru-btn.generating {
    background: var(--yellow-dim);
    border-color: rgba(224,193,69,0.3);
    color: var(--yellow);
  }
  .generate-ru-btn.done {
    background: var(--green-dim);
    border-color: rgba(91,214,159,0.3);
    color: var(--green);
  }

  /* ── Error banner ── */
  .error-banner {
    background: rgba(255,79,218,0.1);
    border: 1px solid rgba(255,79,218,0.25);
    color: #FF9EED;
    padding: 8px 18px;
    font-size: 0.82rem;
    text-align: center;
  }
</style>
</head>
<body>

<!-- Lightbox -->
<div id="lightbox">
  <button id="lightbox-close" title="Close">✕</button>
  <img id="lightbox-img" src="" alt="">
</div>

<!-- Toast -->
<div id="toast"></div>

<header>
  <div class="brand">{{ brand_name }}<span class="brand-dot">.</span></div>
  <div class="header-sep"></div>
  <div class="header-date">{{ date }}</div>
  {% if topics %}
  <span class="header-count">{{ topics|length }} topic{{ 's' if topics|length != 1 else '' }}</span>
  {% endif %}
  <div class="spacer"></div>
  {% if available_dates|length > 1 %}
  <form method="get" action="/" style="margin:0">
    <select class="date-select" name="date" onchange="this.form.submit()" title="Switch date">
      {% for d in available_dates %}
      <option value="{{ d }}"{% if d == date %} selected{% endif %}>{{ d }}</option>
      {% endfor %}
    </select>
  </form>
  {% endif %}
  <a href="/admin" class="admin-link">Admin</a>
  <div class="lang-toggle">
    <button class="lang-btn active" id="btn-en" onclick="setLang('en')">EN</button>
    <button class="lang-btn" id="btn-ru" onclick="setLang('ru')">RU</button>
  </div>
</header>

{% if error %}
<div class="error-banner">{{ error }}</div>
{% endif %}

<main class="grid">
{% if topics %}
  {% for t in topics %}
  <div class="card" id="card-{{ loop.index }}">

    <!-- EN Image -->
    <div class="card-img en-only" {% if t.image_filename %}data-src="/images/{{ t.image_filename }}"{% endif %}
         title="{% if t.image_filename %}Click to enlarge{% endif %}">
      {% if t.image_filename %}
      <img id="img-{{ loop.index }}" src="/images/{{ t.image_filename }}" alt="{{ t.topic }}" loading="lazy">
      <div class="img-overlay"></div>
      {% else %}
      <div class="no-img">
        <span class="no-img-icon">🖼</span>
        <span>No image generated</span>
      </div>
      {% endif %}
      <div class="img-loading-overlay" id="overlay-{{ loop.index }}">
        <div class="spinner"></div>
        <span class="spinner-label">Generating new image…</span>
      </div>
    </div>

    <!-- RU Image -->
    <div class="card-img ru-only" {% if t.image_filename_ru %}data-src="/images/{{ t.image_filename_ru }}"{% endif %}
         title="{% if t.image_filename_ru %}Click to enlarge{% endif %}">
      {% if t.image_filename_ru %}
      <img id="img-ru-{{ loop.index }}" src="/images/{{ t.image_filename_ru }}" alt="{{ t.topic }}" loading="lazy">
      <div class="img-overlay"></div>
      {% else %}
      <div class="no-img" id="no-img-ru-{{ loop.index }}">
        <span class="no-img-icon">🖼</span>
        <span>No Russian image</span>
      </div>
      {% endif %}
      <div class="img-loading-overlay" id="overlay-ru-{{ loop.index }}">
        <div class="spinner"></div>
        <span class="spinner-label">Regenerating RU image…</span>
      </div>
    </div>

    <!-- RU Image actions -->
    <div class="image-actions ru-only">
      <span class="img-status" style="font-size:0.72rem;color:var(--muted)">
        {% if t.image_filename_ru %}RU image{% else %}No RU image{% endif %}
      </span>
      <button class="action-btn regen-btn" id="regen-ru-{{ loop.index }}"
              onclick="regenRuImage({{ loop.index }})"
              {% if not t.image_filename_ru %}disabled{% endif %}>↻ Regenerate</button>
    </div>

    <!-- Image approval actions (EN only) -->
    <div class="image-actions en-only">
      <span class="img-status" id="status-{{ loop.index }}">Pending review</span>
      <button class="action-btn approve-btn" id="approve-{{ loop.index }}"
              onclick="approveImage({{ loop.index }})">✓ Approve</button>
      <button class="action-btn regen-btn" id="regen-{{ loop.index }}"
              onclick="regenImage({{ loop.index }})">↻ Regenerate</button>
    </div>

    <div class="card-body">
      <h2 class="topic-title">{{ t.topic }}</h2>

      <!-- Platform tabs -->
      <div class="tab-bar">
        <button class="tab active" data-tab="twitter-{{ loop.index }}">
          <span class="platform-icon">𝕏</span> Twitter
        </button>
        <button class="tab" data-tab="telegram-{{ loop.index }}">
          <span class="platform-icon">✈</span> Telegram
        </button>
      </div>

      <!-- Twitter panel -->
      <div class="panel active" id="twitter-{{ loop.index }}">
        <!-- EN Twitter -->
        <div class="en-only">
          {% if t.twitter %}
          {% set tweets = t.twitter.split('---') %}
          <div class="content-box" id="tw-box-{{ loop.index }}">{% for tweet in tweets %}{{ tweet.strip() }}{% if not loop.last %}
<hr class="tweet-divider">
{% endif %}{% endfor %}</div>
          <div class="content-actions">
            <span class="char-count">{{ t.twitter|length }} chars</span>
            <button class="copy-btn" data-raw="{{ t.twitter | e }}">Copy</button>
          </div>
          {% else %}
          <div class="content-box" style="color:var(--muted);font-style:italic;">No Twitter content</div>
          {% endif %}
        </div>
        <!-- RU Twitter -->
        <div class="ru-only">
          {% if t.twitter_ru %}
          {% set tweets_ru = t.twitter_ru.split('---') %}
          <div class="content-box" id="tw-box-ru-{{ loop.index }}">{% for tweet in tweets_ru %}{{ tweet.strip() }}{% if not loop.last %}
<hr class="tweet-divider">
{% endif %}{% endfor %}</div>
          <div class="content-actions">
            <span class="char-count">{{ t.twitter_ru|length }} chars</span>
            <button class="copy-btn" data-raw="{{ t.twitter_ru | e }}">Copy</button>
          </div>
          {% else %}
          <div class="content-box" style="color:var(--muted);font-style:italic;">Нет контента для Twitter</div>
          {% endif %}
        </div>
      </div>

      <!-- Telegram panel -->
      <div class="panel" id="telegram-{{ loop.index }}">
        <!-- EN Telegram -->
        <div class="en-only">
          {% if t.telegram %}
          <div class="content-box" id="tg-box-{{ loop.index }}">{{ t.telegram }}</div>
          <div class="content-actions">
            <span class="char-count">{{ t.telegram|length }} chars</span>
            <button class="copy-btn" data-raw="{{ t.telegram | e }}">Copy</button>
          </div>
          {% else %}
          <div class="content-box" style="color:var(--muted);font-style:italic;">No Telegram content</div>
          {% endif %}
        </div>
        <!-- RU Telegram -->
        <div class="ru-only">
          {% if t.telegram_ru %}
          <div class="content-box" id="tg-box-ru-{{ loop.index }}">{{ t.telegram_ru }}</div>
          <div class="content-actions">
            <span class="char-count">{{ t.telegram_ru|length }} chars</span>
            <button class="copy-btn" data-raw="{{ t.telegram_ru | e }}">Copy</button>
          </div>
          {% else %}
          <div class="content-box" style="color:var(--muted);font-style:italic;">Нет контента для Telegram</div>
          {% endif %}
        </div>
      </div>

      <!-- EN Hashtags -->
      {% if t.hashtag_list %}
      <div class="en-only">
        <div class="hashtags">
          {% for tag in t.hashtag_list %}
          <span class="tag" data-tag="{{ tag }}" title="Click to copy">{{ tag }}</span>
          {% endfor %}
        </div>
      </div>
      {% endif %}

      <!-- RU Generate button + Hashtags -->
      <div class="ru-only">
        <div class="generate-ru-bar" id="ru-gen-bar-{{ loop.index }}">
          <span class="ru-gen-label" id="ru-gen-label-{{ loop.index }}">
            {% if t.twitter_ru or t.telegram_ru %}Russian content ready{% else %}No Russian content yet{% endif %}
          </span>
          <button class="generate-ru-btn" id="ru-gen-btn-{{ loop.index }}"
                  onclick="generateRussian({{ loop.index }})"
                  {% if t.twitter_ru and t.telegram_ru %}disabled{% endif %}>
            {% if t.twitter_ru and t.telegram_ru %}✓ Generated{% else %}Generate Russian{% endif %}
          </button>
        </div>
        {% if t.hashtag_list_ru %}
        <div class="hashtags">
          {% for tag in t.hashtag_list_ru %}
          <span class="tag" data-tag="{{ tag }}" title="Click to copy">{{ tag }}</span>
          {% endfor %}
        </div>
        {% endif %}
      </div>

    </div>
  </div>
  {% endfor %}

{% else %}
  <div class="empty">
    <h2>No content found for {{ date }}</h2>
    <p>Run <code>/weekly-pipeline</code> to generate content first.</p>
  </div>
{% endif %}
</main>

<script>
// ── Topic data embedded from server ────────────────────────────────────────
const TOPICS       = {{ topics_json | safe }};
const CURRENT_DATE = {{ date_json | safe }};
const HAS_GENERATOR = {{ has_generator | safe }};
const HAS_RU_GENERATOR = {{ has_ru_generator | safe }};

// ── Language toggle ──────────────────────────────────────────────────────────
function setLang(lang) {
  document.body.classList.toggle('lang-ru', lang === 'ru');
  document.getElementById('btn-en').classList.toggle('active', lang === 'en');
  document.getElementById('btn-ru').classList.toggle('active', lang === 'ru');
  localStorage.setItem('content-dash-lang', lang);
}
// Restore language preference on load
(function(){ if(localStorage.getItem('content-dash-lang')==='ru') setLang('ru'); })();

// ── Toast ───────────────────────────────────────────────────────────────────
let toastTimer;
function showToast(msg, isError = false) {
  const el = document.getElementById('toast');
  el.textContent = msg;
  el.className = 'show' + (isError ? ' error' : '');
  clearTimeout(toastTimer);
  toastTimer = setTimeout(() => { el.className = ''; }, 3000);
}

// ── Tab switching ────────────────────────────────────────────────────────────
document.querySelectorAll('.tab-bar').forEach(bar => {
  bar.querySelectorAll('.tab').forEach(tab => {
    tab.addEventListener('click', () => {
      const card = tab.closest('.card-body');
      const target = tab.dataset.tab;
      bar.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
      card.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
      tab.classList.add('active');
      document.getElementById(target).classList.add('active');
    });
  });
});

// ── Copy buttons ─────────────────────────────────────────────────────────────
function flashCopy(el, successLabel) {
  const orig = el.textContent;
  el.textContent = successLabel;
  el.classList.add('copied');
  setTimeout(() => { el.textContent = orig; el.classList.remove('copied'); }, 2000);
}

function copyText(text) {
  if (navigator.clipboard) {
    return navigator.clipboard.writeText(text);
  }
  const ta = document.createElement('textarea');
  ta.value = text;
  ta.style.position = 'fixed';
  ta.style.opacity = '0';
  document.body.appendChild(ta);
  ta.focus();
  ta.select();
  document.execCommand('copy');
  document.body.removeChild(ta);
  return Promise.resolve();
}

document.querySelectorAll('.copy-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    const text = btn.dataset.raw || '';
    copyText(text).then(() => flashCopy(btn, 'Copied ✓')).catch(() => flashCopy(btn, 'Copied ✓'));
  });
});

// ── Hashtag copy ─────────────────────────────────────────────────────────────
document.querySelectorAll('.tag').forEach(tag => {
  tag.addEventListener('click', () => {
    copyText(tag.dataset.tag).then(() => {
      const orig = tag.textContent;
      tag.textContent = '✓';
      tag.classList.add('tag-copied');
      setTimeout(() => { tag.textContent = orig; tag.classList.remove('tag-copied'); }, 1500);
    });
  });
});

// ── Lightbox ─────────────────────────────────────────────────────────────────
const lightbox = document.getElementById('lightbox');
const lightImg  = document.getElementById('lightbox-img');
const lbClose   = document.getElementById('lightbox-close');

document.querySelectorAll('.card-img[data-src]').forEach(el => {
  el.addEventListener('click', (e) => {
    if (e.target.closest('.img-loading-overlay')) return;
    lightImg.src = el.dataset.src;
    lightbox.classList.add('open');
    document.body.style.overflow = 'hidden';
  });
});

function closeLightbox() {
  lightbox.classList.remove('open');
  document.body.style.overflow = '';
  lightImg.src = '';
}

lbClose.addEventListener('click', closeLightbox);
lightbox.addEventListener('click', e => { if (e.target === lightbox) closeLightbox(); });
document.addEventListener('keydown', e => { if (e.key === 'Escape') closeLightbox(); });

// ── Approval state ───────────────────────────────────────────────────────────
function setApprovalUI(idx, status) {
  const statusEl  = document.getElementById(`status-${idx}`);
  const approveBtn = document.getElementById(`approve-${idx}`);
  const card      = document.getElementById(`card-${idx}`);
  if (!statusEl || !approveBtn) return;

  if (status === 'approved') {
    statusEl.textContent = 'Approved ✓';
    statusEl.className   = 'img-status approved';
    approveBtn.textContent = '✓ Approved';
    approveBtn.classList.add('approved');
    card.classList.add('is-approved');
  } else {
    statusEl.textContent = 'Pending review';
    statusEl.className   = 'img-status';
    approveBtn.textContent = '✓ Approve';
    approveBtn.classList.remove('approved');
    card.classList.remove('is-approved');
  }
}

async function loadApprovals() {
  try {
    const res  = await fetch(`/api/approvals?date=${encodeURIComponent(CURRENT_DATE)}`);
    const data = await res.json();
    TOPICS.forEach((t, i) => {
      const entry = data[t.topic];
      if (entry) {
        setApprovalUI(i + 1, entry.status);
        if (entry.ru_status) t.ru_status = entry.ru_status;
      }
      updateRuGenButton(i + 1);
    });
  } catch (e) {
    console.warn('Could not load approvals:', e);
  }
}

async function approveImage(idx) {
  const t        = TOPICS[idx - 1];
  const approveBtn = document.getElementById(`approve-${idx}`);
  const isApproved = approveBtn.classList.contains('approved');
  const newStatus  = isApproved ? 'pending' : 'approved';

  try {
    await fetch('/api/approve', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ date: CURRENT_DATE, topic: t.topic, status: newStatus }),
    });
    setApprovalUI(idx, newStatus);
    updateRuGenButton(idx);
    showToast(newStatus === 'approved' ? 'Image approved ✓' : 'Approval removed');
  } catch (e) {
    showToast('Could not save approval', true);
  }
}

// ── Regeneration (async + polling) ───────────────────────────────────────────
async function regenImage(idx) {
  if (!HAS_GENERATOR) {
    showToast('Image generator not available, check WAVESPEED_API_KEY', true);
    return;
  }

  const t          = TOPICS[idx - 1];
  const overlay    = document.getElementById(`overlay-${idx}`);
  const approveBtn = document.getElementById(`approve-${idx}`);
  const regenBtn   = document.getElementById(`regen-${idx}`);
  const imgEl      = document.getElementById(`img-${idx}`);
  const cardImg    = imgEl ? imgEl.closest('.card-img') : null;

  overlay.classList.add('active');
  approveBtn.disabled = true;
  regenBtn.disabled   = true;

  try {
    const startRes = await fetch('/api/regenerate', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        date:       CURRENT_DATE,
        topic:      t.topic,
        img_prompt: t.img_prompt || '',
        style:      'tech',
      }),
    });
    const startData = await startRes.json();
    if (!startData.success) {
      showToast('Could not start generation: ' + (startData.error || 'Unknown'), true);
      return;
    }

    const jobId = startData.job_id;
    const result = await pollJob(jobId);

    if (result.status === 'done') {
      if (imgEl) imgEl.src = `/images/${result.filename}?t=${Date.now()}`;
      if (cardImg) cardImg.dataset.src = `/images/${result.filename}`;
      TOPICS[idx - 1].image_filename = result.filename;
      setApprovalUI(idx, 'pending');
      showToast('New image generated — please review');
    } else {
      showToast('Generation failed: ' + (result.error || 'Unknown error'), true);
    }
  } catch (e) {
    showToast('Generation failed: ' + e.message, true);
  } finally {
    overlay.classList.remove('active');
    approveBtn.disabled = false;
    regenBtn.disabled   = false;
  }
}

async function pollJob(jobId, intervalMs = 3000, maxWaitMs = 300000) {
  const deadline = Date.now() + maxWaitMs;
  while (Date.now() < deadline) {
    await new Promise(r => setTimeout(r, intervalMs));
    const res  = await fetch(`/api/regen-status/${jobId}`);
    const data = await res.json();
    if (data.status === 'done' || data.status === 'error') return data;
  }
  return { status: 'error', error: 'Timed out waiting for generation' };
}

// ── Russian generation ───────────────────────────────────────────────────────
function updateRuGenButton(idx) {
  const t = TOPICS[idx - 1];
  const btn = document.getElementById(`ru-gen-btn-${idx}`);
  const label = document.getElementById(`ru-gen-label-${idx}`);
  if (!btn) return;

  // Check if EN is approved
  const approveBtn = document.getElementById(`approve-${idx}`);
  const isApproved = approveBtn && approveBtn.classList.contains('approved');

  // Check if RU content already exists
  const hasRu = t.twitter_ru || t.telegram_ru;

  if (hasRu) {
    btn.textContent = '✓ Generated';
    btn.disabled = true;
    btn.classList.add('done');
    label.textContent = 'Russian content ready';
    label.classList.add('done');
  } else if (!isApproved) {
    btn.textContent = 'Generate Russian';
    btn.disabled = true;
    btn.title = 'Approve EN content first';
    label.textContent = 'Approve English content first';
  } else {
    btn.textContent = 'Generate Russian';
    btn.disabled = false;
    btn.title = '';
    label.textContent = 'Ready to generate Russian content';
  }

  // Check if currently generating (from ru_status)
  if (t.ru_status === 'generating') {
    btn.textContent = 'Generating...';
    btn.disabled = true;
    btn.classList.add('generating');
    label.textContent = 'Generation in progress...';
  }
}

async function generateRussian(idx) {
  if (!HAS_RU_GENERATOR) {
    showToast('Russian generator not available', true);
    return;
  }

  const t = TOPICS[idx - 1];
  const btn = document.getElementById(`ru-gen-btn-${idx}`);
  const label = document.getElementById(`ru-gen-label-${idx}`);

  btn.textContent = 'Translating text...';
  btn.disabled = true;
  btn.classList.add('generating');
  label.textContent = 'Translating text...';

  try {
    const startRes = await fetch('/api/generate-ru', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ date: CURRENT_DATE, topic_index: idx - 1 }),
    });
    const startData = await startRes.json();
    if (!startData.success) {
      showToast('Could not start: ' + (startData.error || 'Unknown'), true);
      btn.textContent = 'Generate Russian';
      btn.disabled = false;
      btn.classList.remove('generating');
      label.textContent = 'Error, try again';
      return;
    }

    const jobId = startData.job_id;
    const result = await pollRuJob(jobId, idx, btn, label);

    if (result.status === 'done') {
      btn.textContent = '✓ Generated';
      btn.classList.remove('generating');
      btn.classList.add('done');
      label.textContent = 'Russian content ready';
      label.classList.add('done');
      showToast('Russian content generated. Reloading...');
      // Reload page to show new RU content
      setTimeout(() => location.reload(), 1500);
    } else {
      btn.textContent = 'Generate Russian';
      btn.disabled = false;
      btn.classList.remove('generating');
      label.textContent = 'Generation failed';
      showToast('Generation failed: ' + (result.error || 'Unknown error'), true);
    }
  } catch (e) {
    btn.textContent = 'Generate Russian';
    btn.disabled = false;
    btn.classList.remove('generating');
    label.textContent = 'Error, try again';
    showToast('Generation failed: ' + e.message, true);
  }
}

async function pollRuJob(jobId, idx, btn, label) {
  const deadline = Date.now() + 300000;
  while (Date.now() < deadline) {
    await new Promise(r => setTimeout(r, 3000));
    const res = await fetch(`/api/ru-status/${jobId}`);
    const data = await res.json();
    // Update phase text
    if (data.phase === 'translating_text') {
      btn.textContent = 'Translating text...';
      label.textContent = 'Translating text via Gemini...';
    } else if (data.phase === 'translating_image') {
      btn.textContent = 'Translating image...';
      label.textContent = 'Translating image via Seedream 4.5...';
    }
    if (data.status === 'done' || data.status === 'error') return data;
  }
  return { status: 'error', error: 'Timed out' };
}

// ── RU image regeneration ─────────────────────────────────────────────────────
async function regenRuImage(idx) {
  if (!HAS_RU_GENERATOR) {
    showToast('RU image generator not available', true);
    return;
  }

  const t = TOPICS[idx - 1];
  const overlay = document.getElementById(`overlay-ru-${idx}`);
  const regenBtn = document.getElementById(`regen-ru-${idx}`);
  const imgEl = document.getElementById(`img-ru-${idx}`);
  const cardImg = imgEl ? imgEl.closest('.card-img') : null;

  if (overlay) overlay.classList.add('active');
  if (regenBtn) regenBtn.disabled = true;

  try {
    const startRes = await fetch('/api/regenerate-ru', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ date: CURRENT_DATE, topic_index: idx - 1 }),
    });
    const startData = await startRes.json();
    if (!startData.success) {
      showToast('Could not start RU regen: ' + (startData.error || 'Unknown'), true);
      return;
    }

    const jobId = startData.job_id;
    const result = await pollJob(jobId);

    if (result.status === 'done' && result.ru_image) {
      const newSrc = `/images/${result.ru_image}?t=${Date.now()}`;
      if (imgEl) {
        imgEl.src = newSrc;
      } else {
        // Replace "no image" placeholder with actual img element
        const noImg = document.getElementById(`no-img-ru-${idx}`);
        if (noImg && noImg.parentNode) {
          const img = document.createElement('img');
          img.id = `img-ru-${idx}`;
          img.src = newSrc;
          img.alt = t.topic;
          img.loading = 'lazy';
          noImg.parentNode.insertBefore(img, noImg);
          const overlay2 = document.createElement('div');
          overlay2.className = 'img-overlay';
          noImg.parentNode.insertBefore(overlay2, noImg);
          noImg.remove();
        }
      }
      if (cardImg) cardImg.dataset.src = `/images/${result.ru_image}`;
      TOPICS[idx - 1].image_filename_ru = result.ru_image;
      showToast('RU image regenerated');
    } else {
      showToast('RU regen failed: ' + (result.error || 'Unknown error'), true);
    }
  } catch (e) {
    showToast('RU regen failed: ' + e.message, true);
  } finally {
    if (overlay) overlay.classList.remove('active');
    if (regenBtn) regenBtn.disabled = false;
  }
}

// ── Init ─────────────────────────────────────────────────────────────────────
loadApprovals();
</script>
</body>
</html>"""


# ── Admin HTML template ────────────────────────────────────────────────────────

ADMIN_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Client Admin</title>
<link rel="icon" type="image/jpeg" href="/favicon.jpg">
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  :root {
    --bg:        #0D1526;
    --surface:   #111B32;
    --surface2:  #162038;
    --border:    rgba(21,137,220,0.15);
    --blue:      #1589DC;
    --blue-dim:  rgba(21,137,220,0.12);
    --green:     #5BD69F;
    --green-dim: rgba(91,214,159,0.12);
    --yellow:    #E0C145;
    --yellow-dim:rgba(224,193,69,0.12);
    --pink:      #FF4FDA;
    --white:     #FFFFFF;
    --muted:     #6B82A8;
    --text:      #C8D8EE;
  }

  html, body {
    background: var(--bg);
    color: var(--white);
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Inter', sans-serif;
    min-height: 100vh;
    line-height: 1.5;
  }

  header {
    background: #080F1E;
    border-bottom: 1px solid var(--border);
    padding: 0 24px;
    height: 60px;
    display: flex;
    align-items: center;
    gap: 16px;
    position: sticky;
    top: 0;
    z-index: 200;
  }

  .brand {
    font-size: 1.1rem;
    font-weight: 700;
    color: var(--blue);
    letter-spacing: 0.3px;
    white-space: nowrap;
  }

  .spacer { flex: 1; }

  .back-link {
    color: var(--muted);
    text-decoration: none;
    font-size: 0.85rem;
    font-weight: 500;
    transition: color 0.15s;
  }
  .back-link:hover { color: var(--blue); }

  .grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(340px, 1fr));
    gap: 24px;
    padding: 28px 24px;
    max-width: 1200px;
    margin: 0 auto;
  }

  @media (max-width: 500px) {
    .grid { grid-template-columns: 1fr; padding: 14px; gap: 16px; }
    header { padding: 0 14px; gap: 10px; }
  }

  .client-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 14px;
    overflow: hidden;
    display: flex;
    flex-direction: column;
    transition: border-color 0.2s, box-shadow 0.2s;
  }
  .client-card:hover {
    border-color: rgba(21,137,220,0.35);
    box-shadow: 0 4px 32px rgba(21,137,220,0.08);
  }
  .client-card.is-active {
    border-color: rgba(91,214,159,0.4);
  }
  .client-card.is-active:hover {
    border-color: rgba(91,214,159,0.6);
    box-shadow: 0 4px 32px rgba(91,214,159,0.08);
  }

  .card-stripe {
    height: 4px;
    width: 100%;
  }

  .card-body {
    padding: 18px 20px 20px;
    display: flex;
    flex-direction: column;
    gap: 14px;
  }

  .card-header {
    display: flex;
    align-items: center;
    gap: 10px;
  }

  .client-name {
    font-size: 1.05rem;
    font-weight: 700;
    color: var(--white);
  }

  .active-badge {
    background: var(--green-dim);
    border: 1px solid rgba(91,214,159,0.3);
    color: var(--green);
    font-size: 0.68rem;
    font-weight: 600;
    padding: 2px 8px;
    border-radius: 20px;
    letter-spacing: 0.5px;
    text-transform: uppercase;
  }

  .client-tagline {
    font-size: 0.82rem;
    color: var(--muted);
    line-height: 1.4;
  }

  .client-website {
    font-size: 0.78rem;
    color: var(--blue);
    text-decoration: none;
    transition: color 0.15s;
  }
  .client-website:hover { color: #1DA0FF; }

  .stats {
    display: flex;
    flex-direction: column;
    gap: 6px;
    padding-top: 4px;
  }
  .stat-row {
    display: flex;
    align-items: center;
    gap: 8px;
    font-size: 0.78rem;
    color: var(--text);
  }
  .stat-label {
    color: var(--muted);
    min-width: 95px;
  }

  .color-swatches {
    display: flex;
    gap: 6px;
    align-items: center;
    padding-top: 2px;
  }
  .swatch {
    width: 20px;
    height: 20px;
    border-radius: 5px;
    border: 1px solid rgba(255,255,255,0.1);
    position: relative;
  }
  .swatch-label {
    font-size: 0.7rem;
    color: var(--muted);
    margin-left: 4px;
  }

  .switch-btn {
    width: 100%;
    padding: 10px;
    border: 1px solid rgba(21,137,220,0.3);
    border-radius: 9px;
    background: var(--blue-dim);
    color: var(--blue);
    font-size: 0.82rem;
    font-weight: 600;
    cursor: pointer;
    transition: all 0.15s;
    margin-top: 4px;
  }
  .switch-btn:hover:not(:disabled) {
    background: rgba(21,137,220,0.2);
    border-color: rgba(21,137,220,0.5);
  }
  .switch-btn:disabled {
    background: var(--green-dim);
    border-color: rgba(91,214,159,0.3);
    color: var(--green);
    cursor: default;
  }

  .new-client-card {
    border-style: dashed;
    border-color: rgba(21,137,220,0.25);
    text-decoration: none;
    cursor: pointer;
  }
  .new-client-card:hover {
    border-color: rgba(21,137,220,0.5);
    background: var(--blue-dim);
  }

  .footer-banner {
    text-align: center;
    padding: 18px 24px;
    font-size: 0.82rem;
    color: var(--muted);
    border-top: 1px solid var(--border);
    margin-top: 12px;
  }
  .footer-banner code {
    background: var(--surface2);
    padding: 2px 7px;
    border-radius: 5px;
    font-size: 0.82rem;
    color: var(--blue);
  }

  #toast {
    position: fixed;
    bottom: 24px;
    left: 50%;
    transform: translateX(-50%) translateY(20px);
    background: #1A2540;
    border: 1px solid var(--border);
    color: var(--text);
    padding: 9px 18px;
    border-radius: 9px;
    font-size: 0.8rem;
    opacity: 0;
    transition: opacity 0.2s, transform 0.2s;
    pointer-events: none;
    z-index: 500;
    white-space: nowrap;
  }
  #toast.show {
    opacity: 1;
    transform: translateX(-50%) translateY(0);
  }
  #toast.error { border-color: rgba(255,79,218,0.3); color: #FF9EED; }
</style>
</head>
<body>

<div id="toast"></div>

<header>
  <div class="brand">Client Admin</div>
  <div class="spacer"></div>
  <a href="/" class="back-link">Back to Dashboard &rarr;</a>
</header>

<main class="grid">
  <a href="/admin/new" class="client-card new-client-card" title="Onboard a new client">
    <div class="card-body" style="align-items:center;justify-content:center;min-height:220px;">
      <span style="font-size:2.5rem;color:var(--muted);opacity:0.5;">+</span>
      <span style="font-size:0.9rem;color:var(--muted);font-weight:600;">New Client</span>
    </div>
  </a>

  {% for c in clients %}
  <div class="client-card{% if c.is_active %} is-active{% endif %}" id="card-{{ c.client_id }}">
    <div class="card-stripe" style="background: {{ c.accent_color }}"></div>
    <div class="card-body">
      <div class="card-header">
        <span class="client-name">{{ c.display_name }}</span>
        {% if c.is_active %}
        <span class="active-badge">Active</span>
        {% endif %}
      </div>

      {% if c.tagline %}
      <div class="client-tagline">{{ c.tagline }}</div>
      {% endif %}

      {% if c.website %}
      <a href="https://{{ c.website }}" target="_blank" rel="noopener" class="client-website">{{ c.website }}</a>
      {% endif %}

      <div class="stats">
        <div class="stat-row">
          <span class="stat-label">Languages</span>
          <span>{{ c.languages | join(', ') | upper }} ({{ c.languages | length }})</span>
        </div>
        <div class="stat-row">
          <span class="stat-label">Keywords</span>
          <span>{{ c.keyword_count }}</span>
        </div>
        <div class="stat-row">
          <span class="stat-label">Content weeks</span>
          <span>{{ c.content_weeks }}</span>
        </div>
      </div>

      <div class="color-swatches">
        <div class="swatch" style="background: {{ c.primary_color }}" title="Primary: {{ c.primary_color }}"></div>
        <div class="swatch" style="background: {{ c.accent_color }}" title="Accent: {{ c.accent_color }}"></div>
        {% if c.secondary_accent %}
        <div class="swatch" style="background: {{ c.secondary_accent }}" title="Secondary: {{ c.secondary_accent }}"></div>
        {% endif %}
        <span class="swatch-label">Brand colors</span>
      </div>

      <button class="switch-btn"
              {% if c.is_active %}disabled{% endif %}
              onclick="switchClient('{{ c.client_id }}')">
        {% if c.is_active %}Active{% else %}Switch to {{ c.display_name }}{% endif %}
      </button>
    </div>
  </div>
  {% endfor %}
</main>

<div class="footer-banner">
  <a href="/admin/new" style="color:var(--blue);text-decoration:none;">+ Onboard a new client</a> or run <code>/onboard-client &lt;name&gt;</code> from the CLI.
</div>

<script>
let toastTimer;
function showToast(msg, isError) {
  const el = document.getElementById('toast');
  el.textContent = msg;
  el.className = 'show' + (isError ? ' error' : '');
  clearTimeout(toastTimer);
  toastTimer = setTimeout(() => { el.className = ''; }, 3000);
}

async function switchClient(clientId) {
  try {
    const res = await fetch('/api/admin/switch-client', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ client_id: clientId }),
    });
    const data = await res.json();
    if (data.success) {
      showToast('Switched to ' + data.display_name);
      setTimeout(() => location.reload(), 800);
    } else {
      showToast(data.error || 'Switch failed', true);
    }
  } catch (e) {
    showToast('Switch failed: ' + e.message, true);
  }
}
</script>
</body>
</html>"""


# ── Onboard HTML template ──────────────────────────────────────────────────────

ONBOARD_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>New Client</title>
<link rel="icon" type="image/jpeg" href="/favicon.jpg">
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  :root {
    --bg:        #0D1526;
    --surface:   #111B32;
    --surface2:  #162038;
    --border:    rgba(21,137,220,0.15);
    --blue:      #1589DC;
    --blue-dim:  rgba(21,137,220,0.12);
    --green:     #5BD69F;
    --green-dim: rgba(91,214,159,0.12);
    --yellow:    #E0C145;
    --pink:      #FF4FDA;
    --white:     #FFFFFF;
    --muted:     #6B82A8;
    --text:      #C8D8EE;
  }

  html, body {
    background: var(--bg);
    color: var(--white);
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Inter', sans-serif;
    min-height: 100vh;
    line-height: 1.5;
  }

  header {
    background: #080F1E;
    border-bottom: 1px solid var(--border);
    padding: 0 24px;
    height: 60px;
    display: flex;
    align-items: center;
    gap: 16px;
    position: sticky;
    top: 0;
    z-index: 200;
  }

  .brand {
    font-size: 1.1rem;
    font-weight: 700;
    color: var(--blue);
    letter-spacing: 0.3px;
    white-space: nowrap;
  }

  .spacer { flex: 1; }

  .back-link {
    color: var(--muted);
    text-decoration: none;
    font-size: 0.85rem;
    font-weight: 500;
    transition: color 0.15s;
  }
  .back-link:hover { color: var(--blue); }

  .form-container {
    max-width: 720px;
    margin: 0 auto;
    padding: 28px 24px 60px;
  }

  .section {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 20px 22px;
    margin-bottom: 20px;
  }

  .section-title {
    font-size: 0.72rem;
    font-weight: 700;
    color: var(--muted);
    letter-spacing: 1px;
    text-transform: uppercase;
    margin-bottom: 16px;
  }

  .field {
    margin-bottom: 14px;
  }
  .field:last-child { margin-bottom: 0; }

  label {
    display: block;
    font-size: 0.8rem;
    font-weight: 600;
    color: var(--text);
    margin-bottom: 5px;
  }

  label .hint {
    font-weight: 400;
    color: var(--muted);
    font-size: 0.75rem;
  }

  input[type="text"],
  input[type="url"],
  textarea {
    width: 100%;
    background: #080F1E;
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 9px 12px;
    color: var(--white);
    font-family: inherit;
    font-size: 0.85rem;
    outline: none;
    transition: border-color 0.15s;
  }
  input:focus, textarea:focus {
    border-color: var(--blue);
  }
  textarea {
    min-height: 80px;
    resize: vertical;
    line-height: 1.5;
  }

  .row-2 {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 14px;
  }

  .color-row {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(130px, 1fr));
    gap: 12px;
  }

  .color-field {
    display: flex;
    flex-direction: column;
    gap: 5px;
  }

  .color-field label {
    font-size: 0.75rem;
  }

  .color-input-wrap {
    display: flex;
    gap: 6px;
    align-items: center;
  }

  input[type="color"] {
    width: 32px;
    height: 32px;
    border: 1px solid var(--border);
    border-radius: 6px;
    background: none;
    cursor: pointer;
    padding: 2px;
  }

  .color-input-wrap input[type="text"] {
    width: 90px;
    font-family: monospace;
    font-size: 0.78rem;
    padding: 6px 8px;
  }

  .check-group {
    display: flex;
    gap: 18px;
    align-items: center;
  }

  .check-group label {
    display: flex;
    align-items: center;
    gap: 6px;
    cursor: pointer;
    font-weight: 500;
  }

  input[type="checkbox"] {
    accent-color: var(--blue);
    width: 16px;
    height: 16px;
  }

  input[type="file"] {
    color: var(--text);
    font-size: 0.82rem;
  }

  .slug-preview {
    font-size: 0.75rem;
    color: var(--muted);
    font-family: monospace;
    margin-top: 3px;
  }

  .submit-btn {
    width: 100%;
    padding: 14px;
    background: var(--blue);
    border: none;
    border-radius: 10px;
    color: #fff;
    font-size: 0.95rem;
    font-weight: 700;
    cursor: pointer;
    transition: background 0.15s, transform 0.1s;
    letter-spacing: 0.3px;
  }
  .submit-btn:hover { background: #1070BB; }
  .submit-btn:active { transform: scale(0.98); }
  .submit-btn:disabled {
    opacity: 0.5;
    cursor: not-allowed;
  }

  #toast {
    position: fixed;
    bottom: 24px;
    left: 50%;
    transform: translateX(-50%) translateY(20px);
    background: #1A2540;
    border: 1px solid var(--border);
    color: var(--text);
    padding: 9px 18px;
    border-radius: 9px;
    font-size: 0.8rem;
    opacity: 0;
    transition: opacity 0.2s, transform 0.2s;
    pointer-events: none;
    z-index: 500;
    white-space: nowrap;
  }
  #toast.show {
    opacity: 1;
    transform: translateX(-50%) translateY(0);
  }
  #toast.error { border-color: rgba(255,79,218,0.3); color: #FF9EED; }

  @media (max-width: 500px) {
    .form-container { padding: 14px; }
    .row-2 { grid-template-columns: 1fr; }
    .color-row { grid-template-columns: 1fr 1fr; }
    header { padding: 0 14px; }
  }
</style>
</head>
<body>

<div id="toast"></div>

<header>
  <div class="brand">New Client</div>
  <div class="spacer"></div>
  <a href="/admin" class="back-link">&larr; Back to Admin</a>
</header>

<div class="form-container">
<form id="onboard-form" enctype="multipart/form-data">

  <!-- Basic Info -->
  <div class="section">
    <div class="section-title">Basic Info</div>
    <div class="field">
      <label>Display Name <span class="hint">(required)</span></label>
      <input type="text" name="display_name" id="display_name" required placeholder="Acme Crypto">
    </div>
    <div class="field">
      <label>Client ID <span class="hint">(auto-generated from name)</span></label>
      <input type="text" name="client_id" id="client_id" placeholder="acme-crypto" readonly>
      <div class="slug-preview" id="slug-preview">clients/<span id="slug-text">...</span>/</div>
    </div>
    <div class="row-2">
      <div class="field">
        <label>Tagline</label>
        <input type="text" name="tagline" placeholder="Short tagline describing the product">
      </div>
      <div class="field">
        <label>Website</label>
        <input type="text" name="website" placeholder="example.com">
      </div>
    </div>
  </div>

  <!-- Brand Colors -->
  <div class="section">
    <div class="section-title">Brand Colors</div>
    <div class="color-row">
      <div class="color-field">
        <label>Primary</label>
        <div class="color-input-wrap">
          <input type="color" name="primary_color_picker" value="#1a1a2e" onchange="syncColor(this, 'primary_color')">
          <input type="text" name="primary_color" value="#1a1a2e" onchange="syncPicker(this, 'primary_color_picker')">
        </div>
      </div>
      <div class="color-field">
        <label>Accent</label>
        <div class="color-input-wrap">
          <input type="color" name="accent_color_picker" value="#1589DC" onchange="syncColor(this, 'accent_color')">
          <input type="text" name="accent_color" value="#1589DC" onchange="syncPicker(this, 'accent_color_picker')">
        </div>
      </div>
      <div class="color-field">
        <label>Text</label>
        <div class="color-input-wrap">
          <input type="color" name="text_color_picker" value="#ffffff" onchange="syncColor(this, 'text_color')">
          <input type="text" name="text_color" value="#ffffff" onchange="syncPicker(this, 'text_color_picker')">
        </div>
      </div>
      <div class="color-field">
        <label>Secondary</label>
        <div class="color-input-wrap">
          <input type="color" name="secondary_accent_picker" value="#FF4FDA" onchange="syncColor(this, 'secondary_accent')">
          <input type="text" name="secondary_accent" value="#FF4FDA" onchange="syncPicker(this, 'secondary_accent_picker')">
        </div>
      </div>
      <div class="color-field">
        <label>Surface</label>
        <div class="color-input-wrap">
          <input type="color" name="surface_color_picker" value="#111B32" onchange="syncColor(this, 'surface_color')">
          <input type="text" name="surface_color" value="#111B32" onchange="syncPicker(this, 'surface_color_picker')">
        </div>
      </div>
    </div>
  </div>

  <!-- Brand Identity -->
  <div class="section">
    <div class="section-title">Brand Identity</div>
    <div class="field">
      <label>Mascot Description <span class="hint">(appearance, style, features)</span></label>
      <textarea name="mascot_description" rows="3" placeholder="A friendly robot mascot with blue eyes and metallic silver body..."></textarea>
    </div>
    <div class="field">
      <label>Logo Description <span class="hint">(how it appears in images)</span></label>
      <input type="text" name="logo_description" placeholder="Centered white logo on dark background, small bottom-right watermark">
    </div>
    <div class="row-2">
      <div class="field">
        <label>Background Style</label>
        <input type="text" name="background_style" placeholder="Dark gradient with subtle particle effects">
      </div>
      <div class="field">
        <label>Background Gradient</label>
        <input type="text" name="background_gradient" placeholder="#111b32 to #070a1b">
      </div>
    </div>
    <div class="field">
      <label>Logo File <span class="hint">(optional, PNG preferred)</span></label>
      <input type="file" name="logo" accept="image/png,image/jpeg,image/webp">
    </div>
  </div>

  <!-- Content Voice -->
  <div class="section">
    <div class="section-title">Content Voice</div>
    <div class="field">
      <label>Tone <span class="hint">(e.g. transparent, educational, no hype)</span></label>
      <input type="text" name="tone" placeholder="Professional, educational, transparent">
    </div>
    <div class="field">
      <label>Voice <span class="hint">(longer description of brand communication style)</span></label>
      <textarea name="voice" rows="3" placeholder="We communicate in a straightforward, educational style. We never overpromise returns..."></textarea>
    </div>
    <div class="row-2">
      <div class="field">
        <label>CTA URL</label>
        <input type="text" name="cta_url" placeholder="example.com">
      </div>
      <div class="field">
        <label>CTA Examples <span class="hint">(one per line)</span></label>
        <textarea name="cta_examples" rows="2" placeholder="Learn more at example.com&#10;Try it free: example.com"></textarea>
      </div>
    </div>
    <div class="field">
      <label>Brand Terms to Keep <span class="hint">(one per line, always capitalized as-is)</span></label>
      <textarea name="brand_terms_keep" rows="2" placeholder="BrandName&#10;ProductName"></textarea>
    </div>
    <div class="field">
      <label>Hashtags <span class="hint">(one per line)</span></label>
      <textarea name="hashtags" rows="2" placeholder="#BrandName&#10;#Crypto"></textarea>
    </div>
    <div class="field">
      <label>Messaging Pillars <span class="hint">(one per line, "Name: Description")</span></label>
      <textarea name="messaging_pillars" rows="3" placeholder="Education: Help users understand DeFi concepts&#10;Trust: Build confidence through transparency"></textarea>
    </div>
  </div>

  <!-- Scraping Config -->
  <div class="section">
    <div class="section-title">Scraping Config</div>
    <div class="field">
      <label>Keywords <span class="hint">(one per line)</span></label>
      <textarea name="keywords" rows="3" placeholder="crypto yield&#10;DeFi automation&#10;trading bot"></textarea>
    </div>
    <div class="field">
      <label>Negative Keywords <span class="hint">(one per line, filter these out)</span></label>
      <textarea name="negative_keywords" rows="2" placeholder="scam&#10;spam&#10;rug pull"></textarea>
    </div>
    <div class="field">
      <label>Subreddits <span class="hint">(one per line, without r/)</span></label>
      <textarea name="subreddits" rows="2" placeholder="cryptocurrency&#10;defi"></textarea>
    </div>
  </div>

  <!-- Image Style Presets -->
  <div class="section">
    <div class="section-title">Image Style Presets</div>
    <div class="field">
      <label>Minimal</label>
      <textarea name="preset_minimal" rows="2">clean minimalist banner, solid dark background</textarea>
    </div>
    <div class="field">
      <label>Tech</label>
      <textarea name="preset_tech" rows="2">futuristic data visualization banner with glowing elements</textarea>
    </div>
    <div class="field">
      <label>Notification</label>
      <textarea name="preset_notification" rows="2">realistic smartphone mockup showing app notification</textarea>
    </div>
  </div>

  <!-- Languages -->
  <div class="section">
    <div class="section-title">Languages</div>
    <div class="check-group">
      <label><input type="checkbox" name="lang_en" checked> English</label>
      <label><input type="checkbox" name="lang_ru"> Russian</label>
    </div>
  </div>

  <button type="submit" class="submit-btn" id="submit-btn">Create Client</button>

</form>
</div>

<script>
function slugify(text) {
  return text.toLowerCase().replace(/[^a-z0-9]+/g, '-').replace(/^-|-$/g, '').substring(0, 40);
}

const nameInput = document.getElementById('display_name');
const idInput = document.getElementById('client_id');
const slugText = document.getElementById('slug-text');

nameInput.addEventListener('input', () => {
  const slug = slugify(nameInput.value);
  idInput.value = slug;
  slugText.textContent = slug || '...';
});

function syncColor(picker, textName) {
  document.querySelector(`input[name="${textName}"]`).value = picker.value.toUpperCase();
}
function syncPicker(text, pickerName) {
  const picker = document.querySelector(`input[name="${pickerName}"]`);
  if (/^#[0-9A-Fa-f]{6}$/.test(text.value)) picker.value = text.value;
}

function linesArray(name) {
  const el = document.querySelector(`[name="${name}"]`);
  if (!el) return [];
  return el.value.split('\\n').map(s => s.trim()).filter(Boolean);
}

let toastTimer;
function showToast(msg, isError) {
  const el = document.getElementById('toast');
  el.textContent = msg;
  el.className = 'show' + (isError ? ' error' : '');
  clearTimeout(toastTimer);
  toastTimer = setTimeout(() => { el.className = ''; }, 4000);
}

document.getElementById('onboard-form').addEventListener('submit', async (e) => {
  e.preventDefault();
  const btn = document.getElementById('submit-btn');
  btn.disabled = true;
  btn.textContent = 'Creating...';

  try {
    const form = e.target;
    const fd = new FormData(form);

    // Add checkbox values explicitly
    fd.set('lang_en', form.lang_en.checked ? '1' : '0');
    fd.set('lang_ru', form.lang_ru.checked ? '1' : '0');

    const res = await fetch('/api/admin/create-client', {
      method: 'POST',
      body: fd,
    });
    const data = await res.json();

    if (data.success) {
      showToast('Client created: ' + data.client_id);
      setTimeout(() => { window.location.href = '/admin'; }, 1200);
    } else {
      showToast(data.error || 'Creation failed', true);
      btn.disabled = false;
      btn.textContent = 'Create Client';
    }
  } catch (err) {
    showToast('Error: ' + err.message, true);
    btn.disabled = false;
    btn.textContent = 'Create Client';
  }
});
</script>
</body>
</html>"""


# ── Approval API routes ────────────────────────────────────────────────────────

@app.route("/api/approvals")
def api_get_approvals():
    date = request.args.get("date", "")
    return jsonify(load_approvals(date))


@app.route("/api/approve", methods=["POST"])
def api_approve():
    data   = request.get_json(force=True)
    date   = data.get("date", "")
    topic  = data.get("topic", "")
    status = data.get("status", "approved")

    if not date or not topic:
        return jsonify({"success": False, "error": "date and topic required"}), 400

    approvals = load_approvals(date)
    approvals[topic] = {"status": status}
    save_approvals(date, approvals)
    return jsonify({"success": True})


@app.route("/api/regenerate", methods=["POST"])
def api_regenerate():
    """Start an async EN image regeneration job (Gemini). Returns a job_id immediately."""
    if not HAS_GENERATOR:
        return jsonify({"success": False, "error": "Image generator not available. Check WAVESPEED_API_KEY and imports."}), 503

    data       = request.get_json(force=True)
    date       = data.get("date", "")
    topic      = data.get("topic", "")
    img_prompt = data.get("img_prompt", "")
    style      = data.get("style", "tech")

    if not date or not topic:
        return jsonify({"success": False, "error": "date and topic required"}), 400

    topic_slug   = re.sub(r"[^a-z0-9]+", "_", topic.lower())[:30].strip("_")
    timestamp    = int(time.time())
    date_slug    = date.replace("week:", "")
    subdir       = f"{date_slug}-weekly"
    filename     = f"{subdir}/{date_slug}_{topic_slug}_regen_{timestamp}.png"
    regen_dir    = IMAGES_DIR / subdir
    regen_dir.mkdir(parents=True, exist_ok=True)
    output_path  = str(IMAGES_DIR / filename)
    job_id       = str(uuid.uuid4())

    with _jobs_lock:
        _jobs[job_id] = {"status": "running", "filename": None, "error": None}

    def _run():
        try:
            prompt = img_prompt if img_prompt else build_prompt(topic=topic, style=style)
            gen_img(prompt, output_path, get_default_references(_active_client))
            approvals = load_approvals(date)
            approvals[topic] = {"status": "pending", "image": filename}
            save_approvals(date, approvals)

            # Persist new image path to workbook so it survives page reloads
            xlsx_path = find_excel(date)
            if xlsx_path:
                relative_path = f"outputs/content/images/{filename}"
                wb = openpyxl.load_workbook(str(xlsx_path))
                ws = wb["Content"]
                for row in ws.iter_rows(min_row=2):
                    if row[2].value == topic:  # col C = Topic
                        row[7].value = relative_path  # col H = Image_Path
                wb.save(str(xlsx_path))

            with _jobs_lock:
                _jobs[job_id] = {"status": "done", "filename": filename, "error": None}
        except Exception as e:
            with _jobs_lock:
                _jobs[job_id] = {"status": "error", "filename": None, "error": str(e)}

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"success": True, "job_id": job_id})


@app.route("/api/regen-status/<job_id>")
def api_regen_status(job_id):
    """Poll for job completion."""
    with _jobs_lock:
        job = _jobs.get(job_id)
    if job is None:
        return jsonify({"status": "unknown"}), 404
    return jsonify(job)


# ── Russian generation API ─────────────────────────────────────────────────

@app.route("/api/generate-ru", methods=["POST"])
def api_generate_ru():
    """Start async Russian content generation for a topic. Requires EN to be approved."""
    if not HAS_RU_GENERATOR:
        return jsonify({"success": False, "error": "Russian generator not available. Check imports."}), 503

    data = request.get_json(force=True)
    date = data.get("date", "")
    topic_index = data.get("topic_index")  # 0-based index into topics list

    if not date or topic_index is None:
        return jsonify({"success": False, "error": "date and topic_index required"}), 400

    # Validate EN is approved
    approvals = load_approvals(date)
    xlsx = find_excel(date)
    if not xlsx:
        return jsonify({"success": False, "error": f"No workbook found for {date}"}), 404

    topics = load_content(xlsx)
    if topic_index < 0 or topic_index >= len(topics):
        return jsonify({"success": False, "error": "Invalid topic_index"}), 400

    topic_data = topics[topic_index]
    topic_name = topic_data["topic"]

    approval_entry = approvals.get(topic_name, {})
    if approval_entry.get("status") != "approved":
        return jsonify({"success": False, "error": "English content must be approved first"}), 400

    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = {"status": "running", "phase": "translating_text", "error": None}

    # Update approval state to show generating
    approvals[topic_name]["ru_status"] = "generating"
    save_approvals(date, approvals)

    def _run_ru_generation():
        try:
            week_of = date.replace("week:", "")
            xlsx_path = find_excel(date)

            # Find the workbook row indices for this topic (Twitter + Telegram rows)
            wb_data = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
            ws = wb_data["Content"]
            topic_rows = []  # list of (row_index_1based, platform)
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                if row and row[2] == topic_name:
                    platform = (row[3] or "").lower()
                    topic_rows.append((idx, platform))
            wb_data.close()

            if not topic_rows:
                raise RuntimeError(f"Topic not found in workbook: {topic_name}")

            # Phase 1: Translate text
            with _jobs_lock:
                _jobs[job_id]["phase"] = "translating_text"

            for row_idx, platform in topic_rows:
                plat_key = "twitter" if "twitter" in platform else "telegram"
                en_content = topic_data.get(plat_key) or ""
                if not en_content:
                    continue

                content_ru = translate_text_to_russian(en_content, platform=plat_key)

                # Translate hashtags
                en_hashtags = topic_data.get("hashtags", "")
                hashtags_ru = translate_hashtags_to_russian(str(en_hashtags)) if en_hashtags else ""

                # Update workbook (leave image columns empty for now)
                update_ru_columns(
                    week_of, row_idx,
                    content_ru=content_ru,
                    image_prompt_ru="",
                    image_path_ru="",
                    hashtags_ru=hashtags_ru,
                )

            # Phase 2: Translate image
            with _jobs_lock:
                _jobs[job_id]["phase"] = "translating_image"

            en_image = topic_data.get("image_filename")
            ru_image_filename = None
            if en_image:
                en_image_path = IMAGES_DIR / en_image
                if en_image_path.exists():
                    # Build RU image path: insert _ru before .png
                    stem = Path(en_image).stem
                    parent = Path(en_image).parent
                    ru_image_filename = str(parent / f"{stem}_ru.png")
                    ru_output_path = IMAGES_DIR / ru_image_filename

                    wavespeed_translate_image(str(en_image_path), str(ru_output_path))

                    # Update image path in workbook for first row of this topic
                    if topic_rows:
                        first_row_idx = topic_rows[0][0]
                        relative_path = f"outputs/content/images/{ru_image_filename}"
                        wb2 = openpyxl.load_workbook(str(xlsx_path))
                        ws2 = wb2["Content"]
                        ws2.cell(row=first_row_idx + 1, column=12, value=relative_path)
                        wb2.save(str(xlsx_path))

            # Update approval state
            approvals_updated = load_approvals(date)
            approvals_updated[topic_name]["ru_status"] = "done"
            save_approvals(date, approvals_updated)

            with _jobs_lock:
                _jobs[job_id] = {
                    "status": "done",
                    "phase": "done",
                    "error": None,
                    "ru_image": ru_image_filename,
                }

        except Exception as e:
            approvals_err = load_approvals(date)
            if topic_name in approvals_err:
                approvals_err[topic_name]["ru_status"] = "error"
                save_approvals(date, approvals_err)
            with _jobs_lock:
                _jobs[job_id] = {"status": "error", "phase": "error", "error": str(e)}

    threading.Thread(target=_run_ru_generation, daemon=True).start()
    return jsonify({"success": True, "job_id": job_id})


@app.route("/api/ru-status/<job_id>")
def api_ru_status(job_id):
    """Poll for Russian generation job completion."""
    with _jobs_lock:
        job = _jobs.get(job_id)
    if job is None:
        return jsonify({"status": "unknown"}), 404
    return jsonify(job)


# ── RU image regeneration API ──────────────────────────────────────────────

@app.route("/api/regenerate-ru", methods=["POST"])
def api_regenerate_ru():
    """Re-translate the EN image to RU via WaveSpeed Seedream 4.5."""
    if not HAS_RU_GENERATOR:
        return jsonify({"success": False, "error": "RU image generator not available. Check wavespeed_img import."}), 503

    data = request.get_json(force=True)
    date = data.get("date", "")
    topic_index = data.get("topic_index")  # 0-based

    if not date or topic_index is None:
        return jsonify({"success": False, "error": "date and topic_index required"}), 400

    xlsx = find_excel(date)
    if not xlsx:
        return jsonify({"success": False, "error": f"No workbook found for {date}"}), 404

    topics = load_content(xlsx)
    if topic_index < 0 or topic_index >= len(topics):
        return jsonify({"success": False, "error": "Invalid topic_index"}), 400

    topic_data = topics[topic_index]
    topic_name = topic_data["topic"]
    en_image = topic_data.get("image_filename")

    if not en_image:
        return jsonify({"success": False, "error": "No EN image to translate from"}), 400

    en_image_path = IMAGES_DIR / en_image
    if not en_image_path.exists():
        return jsonify({"success": False, "error": f"EN image not found: {en_image}"}), 404

    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = {"status": "running", "phase": "translating_image", "error": None}

    def _run_ru_regen():
        try:
            # Build RU image path: insert _ru before .png (remove old _ru suffix if re-regenerating)
            stem = Path(en_image).stem
            parent = Path(en_image).parent
            # Strip existing _ru suffix to avoid _ru_ru
            base_stem = re.sub(r"_ru$", "", stem)
            ru_image_filename = str(parent / f"{base_stem}_ru.png")
            ru_output_path = IMAGES_DIR / ru_image_filename

            wavespeed_translate_image(str(en_image_path), str(ru_output_path))

            # Update workbook col L (Image_Path_RU) for rows matching this topic
            xlsx_path = find_excel(date)
            if xlsx_path:
                relative_path = f"outputs/content/images/{ru_image_filename}"
                wb = openpyxl.load_workbook(str(xlsx_path))
                ws = wb["Content"]
                for row in ws.iter_rows(min_row=2):
                    if row[2].value == topic_name:  # col C = Topic
                        row[11].value = relative_path  # col L = Image_Path_RU
                wb.save(str(xlsx_path))

            with _jobs_lock:
                _jobs[job_id] = {
                    "status": "done",
                    "phase": "done",
                    "error": None,
                    "ru_image": ru_image_filename,
                }

        except Exception as e:
            with _jobs_lock:
                _jobs[job_id] = {"status": "error", "phase": "error", "error": str(e)}

    threading.Thread(target=_run_ru_regen, daemon=True).start()
    return jsonify({"success": True, "job_id": job_id})


# ── Admin routes ───────────────────────────────────────────────────────────────

@app.route("/admin/new")
def admin_new_client():
    return render_template_string(ONBOARD_HTML)


@app.route("/api/admin/create-client", methods=["POST"])
def api_admin_create_client():
    """Create a new client from the onboarding form (multipart/form-data)."""
    import shutil

    try:
        display_name = (request.form.get("display_name") or "").strip()
        if not display_name:
            return jsonify({"success": False, "error": "Display name is required"}), 400

        # Slugify client_id
        raw_id = (request.form.get("client_id") or "").strip()
        if not raw_id:
            raw_id = re.sub(r"[^a-z0-9]+", "-", display_name.lower()).strip("-")[:40]
        client_id = raw_id

        if not client_id:
            return jsonify({"success": False, "error": "Could not generate a valid client ID"}), 400

        # Check for existing client
        client_dir = PROJECT_ROOT / "clients" / client_id
        if client_dir.exists():
            return jsonify({"success": False, "error": f"Client '{client_id}' already exists"}), 409

        # Gather form fields
        tagline = (request.form.get("tagline") or "").strip()
        website = (request.form.get("website") or "").strip()
        primary_color = (request.form.get("primary_color") or "#1a1a2e").strip()
        accent_color = (request.form.get("accent_color") or "#1589DC").strip()
        text_color = (request.form.get("text_color") or "#ffffff").strip()
        secondary_accent = (request.form.get("secondary_accent") or "#FF4FDA").strip()
        surface_color = (request.form.get("surface_color") or "#111B32").strip()
        background_gradient = (request.form.get("background_gradient") or f"{surface_color} to #070a1b").strip()
        mascot_description = (request.form.get("mascot_description") or "").strip()
        logo_description = (request.form.get("logo_description") or "").strip()
        background_style = (request.form.get("background_style") or "").strip()
        tone = (request.form.get("tone") or "").strip()
        voice = (request.form.get("voice") or "").strip()
        cta_url = (request.form.get("cta_url") or website).strip()

        def _lines(field_name):
            val = (request.form.get(field_name) or "").strip()
            return [line.strip() for line in val.split("\n") if line.strip()] if val else []

        cta_examples = _lines("cta_examples")
        brand_terms = _lines("brand_terms_keep")
        hashtags = _lines("hashtags")
        pillars = _lines("messaging_pillars")
        keywords = _lines("keywords")
        negative_keywords = _lines("negative_keywords")
        subreddits = _lines("subreddits")
        preset_minimal = (request.form.get("preset_minimal") or "clean minimalist banner, solid dark background").strip()
        preset_tech = (request.form.get("preset_tech") or "futuristic data visualization banner with glowing elements").strip()
        preset_notification = (request.form.get("preset_notification") or "realistic smartphone mockup showing app notification").strip()

        languages = []
        if request.form.get("lang_en") == "1":
            languages.append("en")
        if request.form.get("lang_ru") == "1":
            languages.append("ru")
        if not languages:
            languages = ["en"]

        # Create directory structure
        brand_dir = client_dir / "brand"
        brand_dir.mkdir(parents=True, exist_ok=True)

        # Copy brand README from template
        template_brand_readme = PROJECT_ROOT / "clients" / "_template" / "brand" / "README.md"
        if template_brand_readme.exists():
            shutil.copy2(str(template_brand_readme), str(brand_dir / "README.md"))

        # Save logo if uploaded
        logo_file = request.files.get("logo")
        has_logo = False
        if logo_file and logo_file.filename:
            logo_file.save(str(brand_dir / "logo.png"))
            has_logo = True

        # Build config.json
        config = {
            "client_id": client_id,
            "display_name": display_name,
            "tagline": tagline,
            "website": website,
            "brand": {
                "primary_color": primary_color,
                "accent_color": accent_color,
                "text_color": text_color,
                "secondary_accent": secondary_accent,
                "surface_color": surface_color,
                "background_gradient": background_gradient,
                "logo_path": "brand/logo.png",
                "reference_images": ["brand/logo.png"] if has_logo else [],
                "mascot_description": mascot_description,
                "logo_description": logo_description,
                "background_style": background_style,
            },
            "content": {
                "tone": tone,
                "voice": voice,
                "brand_terms_keep": brand_terms if brand_terms else [display_name],
                "cta_url": cta_url,
                "cta_examples": cta_examples,
                "hashtags": hashtags,
                "messaging_pillars": pillars,
            },
            "scraping": {
                "keywords": keywords,
                "negative_keywords": negative_keywords if negative_keywords else ["scam", "spam"],
                "subreddits": subreddits,
            },
            "image": {
                "style_presets": {
                    "minimal": preset_minimal,
                    "tech": preset_tech,
                    "notification": preset_notification,
                }
            },
            "languages": languages,
        }

        (client_dir / "config.json").write_text(json.dumps(config, indent=2) + "\n")

        # Generate content-guidelines.md
        pillar_sections = ""
        for i, p in enumerate(pillars, 1):
            parts = p.split(":", 1)
            name = parts[0].strip()
            desc = parts[1].strip() if len(parts) > 1 else ""
            pillar_sections += f"\n### {i}. {name}\n{desc}\n"

        hashtag_str = ", ".join(hashtags) if hashtags else "[No hashtags configured]"
        cta_str = "\n".join(f"- {c}" for c in cta_examples) if cta_examples else "- [No CTAs configured]"

        guidelines = f"""# Content Guidelines

Reference for content generation. All generated content must align with these principles.

---

## Brand Voice

**{tone if tone else '[Not configured]'}**

{voice if voice else '[Voice description not provided. Fill in details about how the brand communicates.]'}

---

## Messaging Pillars

Every piece of content should connect to one or more of these:
{pillar_sections if pillar_sections else '''
### 1. [Pillar Name]
[Description]
'''}
---

## Twitter / X Guidelines

### Single Post (280 chars max)
- Concise, value-driven statement
- Include 2-3 relevant hashtags

### Thread (3-5 tweets)
- Tweet 1: Hook
- Tweets 2-3: Value/education
- Tweet 4: Brand connection
- Tweet 5: CTA

---

## Telegram Guidelines

- 300-600 characters per post
- More detailed than Twitter, educational tone
- End with a clear CTA

---

## What to Always Avoid

| Avoid | Why |
|-------|-----|
| Em-dashes | Brand style preference |
| Guaranteed returns | Regulatory compliance |
| Hype language | Brand credibility |

---

## CTAs by Platform

**Twitter:**
{cta_str}

**Telegram:**
{cta_str}

---

## Hashtag Library

**Always relevant:** {hashtag_str}
"""
        (client_dir / "content-guidelines.md").write_text(guidelines)

        # Generate keywords.md
        primary_kw = "\n".join(f"- {k}" for k in keywords) if keywords else "- [Add keywords]"
        neg_kw = "\n".join(f"- {k}" for k in negative_keywords) if negative_keywords else "- scam\n- spam"
        sub_list = "\n".join(f"- r/{s}" for s in subreddits) if subreddits else "- r/[subreddit]"

        keywords_md = f"""# Keyword Reference

Used by the content pipeline to filter and rank scraped posts for relevance.

---

## Primary Keywords (High Relevance)

{primary_kw}

---

## Negative Keywords (Filter Out)

{neg_kw}

---

## Subreddits to Monitor

{sub_list}
"""
        (client_dir / "keywords.md").write_text(keywords_md)

        # Generate context.md
        context_md = f"""# {display_name} — Client Context

## Organization Overview

{tagline if tagline else '[Describe the organization, what it does, and its mission.]'}

Website: {website if website else '[Not provided]'}

---

## Products / Services

* [Add products/services]

---

## Target Audience (ICP)

* [Demographics]
* [Pain points]
* [Goals]

---

## Positioning

* [How the brand differentiates]
* [Key competitive advantages]
"""
        (client_dir / "context.md").write_text(context_md)

        # Create output directory
        output_dir = PROJECT_ROOT / "outputs" / "content" / client_id
        output_dir.mkdir(parents=True, exist_ok=True)

        return jsonify({"success": True, "client_id": client_id})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


def _gather_client_summaries():
    """Build a list of summary dicts for all configured clients."""
    active = client_config.get_active_client()
    summaries = []
    for cid in client_config.list_clients():
        try:
            cfg = client_config.load_config(cid)
        except Exception:
            continue
        brand = cfg.get("brand", {})
        # Count content weeks (xlsx files)
        out_dir = client_config.get_output_dir(cid)
        content_weeks = len(list(out_dir.glob("*-weekly-content.xlsx"))) if out_dir.exists() else 0
        summaries.append({
            "client_id":       cid,
            "display_name":    cfg.get("display_name", cid),
            "tagline":         cfg.get("tagline", ""),
            "website":         cfg.get("website", ""),
            "languages":       cfg.get("languages", ["en"]),
            "keyword_count":   len(cfg.get("scraping", {}).get("keywords", [])),
            "content_weeks":   content_weeks,
            "primary_color":   brand.get("primary_color", "#1a1a2e"),
            "accent_color":    brand.get("accent_color", "#1589DC"),
            "secondary_accent": brand.get("secondary_accent", ""),
            "is_active":       cid == active,
        })
    return summaries


@app.route("/admin")
def admin_page():
    return render_template_string(ADMIN_HTML, clients=_gather_client_summaries())


@app.route("/api/admin/switch-client", methods=["POST"])
def api_admin_switch():
    data = request.get_json(force=True)
    cid = data.get("client_id", "").strip()
    if not cid:
        return jsonify({"success": False, "error": "client_id required"}), 400
    available = client_config.list_clients()
    if cid not in available:
        return jsonify({"success": False, "error": f"Unknown client: {cid}"}), 404
    _reload_active_client(cid)
    return jsonify({"success": True, "client_id": cid, "display_name": _display_name})


@app.route("/api/admin/clients")
def api_admin_clients():
    return jsonify(_gather_client_summaries())


# ── Flask routes ───────────────────────────────────────────────────────────────

@app.route("/favicon.jpg")
def serve_favicon():
    return send_from_directory(str(PROJECT_ROOT / "reference"), "favicon.jpg")


@app.route("/images/<path:filename>")
def serve_image(filename):
    if (IMAGES_DIR / filename).exists():
        return send_from_directory(str(IMAGES_DIR), filename)
    return send_from_directory(str(CONTENT_DIR), filename)


@app.route("/")
def index():
    date_param      = request.args.get("date")
    available_dates = list_available_dates()
    error           = None

    if not available_dates:
        return render_template_string(
            HTML,
            topics=[], date="—", available_dates=[],
            topics_json="[]", date_json='""',
            has_generator="true" if HAS_GENERATOR else "false",
            has_ru_generator="true" if HAS_RU_GENERATOR else "false",
            brand_name=_display_name,
            error=f"No weekly content Excel files found for {_display_name}",
        )

    selected = date_param if date_param in available_dates else available_dates[0]
    xlsx     = find_excel(selected)

    if not xlsx:
        error  = f"No Excel file found for {selected}"
        topics = []
    else:
        topics = load_content(xlsx)

    # Load approvals to embed ru_status in topic data
    current_approvals = load_approvals(selected)

    topics_json = json.dumps([
        {
            "topic":             t["topic"],
            "img_prompt":        t.get("img_prompt", ""),
            "img_prompt_ru":     t.get("img_prompt_ru", ""),
            "image_filename":    t.get("image_filename", ""),
            "image_filename_ru": t.get("image_filename_ru", ""),
            "date":              t.get("date", ""),
            "twitter":           t.get("twitter", ""),
            "telegram":          t.get("telegram", ""),
            "hashtags":          t.get("hashtags", ""),
            "twitter_ru":        t.get("twitter_ru", ""),
            "telegram_ru":       t.get("telegram_ru", ""),
            "hashtags_ru":       t.get("hashtags_ru", ""),
            "ru_status":         current_approvals.get(t["topic"], {}).get("ru_status", ""),
        }
        for t in topics
    ])

    return render_template_string(
        HTML,
        topics=topics,
        date=selected,
        available_dates=available_dates,
        topics_json=topics_json,
        date_json=json.dumps(selected),
        has_generator="true" if HAS_GENERATOR else "false",
        has_ru_generator="true" if HAS_RU_GENERATOR else "false",
        brand_name=_display_name,
        error=error,
    )


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Support --client flag from command line
    import argparse as _ap
    _parser = _ap.ArgumentParser(add_help=False)
    _parser.add_argument("--client", default=None)
    _parser.add_argument("date", nargs="?", default=None)
    _args, _ = _parser.parse_known_args()
    if _args.client:
        _reload_active_client(_args.client)

    print(f"\n  {_display_name} Content Dashboard")
    print(f"  Client      : {_active_client}")
    print(f"  Content dir : {CONTENT_DIR}")
    dates = list_available_dates()
    print(f"  Found dates : {', '.join(dates) if dates else 'none'}")
    print(f"  Generator   : {'available' if HAS_GENERATOR else 'unavailable (check WAVESPEED_API_KEY)'}")
    print(f"\n  → http://localhost:5001\n")
    print("  Press Ctrl+C to stop.\n")
    app.run(host="127.0.0.1", port=5001, debug=False, threaded=True)
