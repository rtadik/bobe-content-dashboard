#!/usr/bin/env python3
"""
X (Twitter) Publisher

Publishes Twitter threads from the weekly content workbook to X (Twitter).
Uses OAuth 1.0a credentials stored in .env as {CLIENT_ID_UPPER}_X_API_KEY etc.

Duplicate prevention: if the topic's Status column is already "Published", exits 0.
Output protocol: prints TWEET_URLS:[...] as last line for Flask subprocess parser.

Usage:
    python scripts/x_publisher.py --client bobe --week-of 2026-02-16 --topic-index 0
    python scripts/x_publisher.py --client bobe --week-of 2026-02-16 --topic-index 0 --mock
"""

import sys
import json
import argparse
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_ROOT / ".env")
except ImportError:
    pass

import client_config

try:
    import tweepy
except ImportError:
    print("tweepy not installed. Run: venv/bin/pip install tweepy", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: venv/bin/pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DAYS_SET = {"Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"}


def load_x_client(client_id: str) -> tweepy.Client:
    """Load tweepy.Client using OAuth 1.0a credentials from environment."""
    credentials = client_config.get_x_credentials(client_id)
    return tweepy.Client(
        consumer_key=credentials["api_key"],
        consumer_secret=credentials["api_secret"],
        access_token=credentials["access_token"],
        access_token_secret=credentials["access_token_secret"],
    )


def split_thread(content: str) -> list:
    """Split thread content on '---' separator into individual tweets."""
    tweets = [t.strip() for t in content.split("---") if t.strip()]
    for i, tweet in enumerate(tweets):
        if len(tweet) > 280:
            print(f"  WARNING: Tweet {i+1} is {len(tweet)} chars (max 280)")
    return tweets


def post_thread(client, tweets: list, mock: bool = False) -> list:
    """Post a Twitter thread sequentially. Returns list of tweet URLs."""
    if mock:
        print("  [MOCK] Would post thread:")
        urls = []
        for i, tweet in enumerate(tweets):
            print(f"  Tweet {i+1} ({len(tweet)} chars): {tweet[:80]}{'...' if len(tweet) > 80 else ''}")
            urls.append(f"mock://tweet/{i+1}")
        return urls

    urls = []
    reply_to_id = None

    for tweet in tweets:
        try:
            if reply_to_id:
                response = client.create_tweet(text=tweet, in_reply_to_tweet_id=reply_to_id)
            else:
                response = client.create_tweet(text=tweet)
            tweet_id = response.data["id"]
            reply_to_id = tweet_id
            url = f"https://twitter.com/i/web/status/{tweet_id}"
            urls.append(url)
            print(f"  Posted tweet {len(urls)}: {url}")
        except tweepy.TooManyRequests:
            import time
            print("  Rate limited. Waiting 60 seconds and retrying...")
            time.sleep(60)
            if reply_to_id:
                response = client.create_tweet(text=tweet, in_reply_to_tweet_id=reply_to_id)
            else:
                response = client.create_tweet(text=tweet)
            tweet_id = response.data["id"]
            reply_to_id = tweet_id
            url = f"https://twitter.com/i/web/status/{tweet_id}"
            urls.append(url)
            print(f"  Posted tweet {len(urls)}: {url}")

    return urls


def _discover_columns(ws):
    """Read header row and return a name->0-based-index map, or None if no headers found."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return None
    col_map = {}
    for idx, val in enumerate(header_row):
        if val:
            col_map[str(val).strip()] = idx
    required = ["Topic", "Platform", "Content", "Status"]
    missing = [h for h in required if h not in col_map]
    if missing:
        print(f"  Warning: Excel headers missing columns: {', '.join(missing)}")
        print(f"  Found: {list(col_map.keys())}")
        return None
    return col_map


def read_twitter_rows(excel_path: Path, topic_index: int):
    """
    Find Excel rows for the twitter content of the given topic_index (0-based).
    Returns (rows, topic_name) where rows is a list of dicts:
      { row_num, platform, content, status, tweet_url }
    """
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb["Content"]

    # Try header-based column discovery first
    col_map = _discover_columns(ws)

    topic_order = []
    topic_rows = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 2:
            continue

        if col_map:
            # Header-validated column positions
            topic_name = row[col_map["Topic"]] if "Topic" in col_map else None
            platform = (row[col_map["Platform"]] or "").lower() if "Platform" in col_map else ""
            content = row[col_map["Content"]] if "Content" in col_map else ""
            status = row[col_map["Status"]] if "Status" in col_map and col_map["Status"] < len(row) else "Draft"
            tweet_url = row[col_map.get("Tweet_URL", 15)] if col_map.get("Tweet_URL", 15) < len(row) else ""
        else:
            # Fallback: positional detection
            col1 = row[1] if len(row) > 1 else None
            if not col1:
                continue

            if col1 in DAYS_SET:
                topic_name = row[2] if len(row) > 2 else None
                platform = (row[3] or "").lower() if len(row) > 3 else ""
                content = row[5] if len(row) > 5 else ""
                status = "Draft"
                tweet_url = ""
            else:
                topic_name = row[3] if len(row) > 3 else None
                platform = (row[4] or "").lower() if len(row) > 4 else ""
                content = row[6] if len(row) > 6 else ""
                status = row[14] if len(row) > 14 else "Draft"
                tweet_url = row[15] if len(row) > 15 else ""

        if not topic_name:
            continue

        if topic_name not in topic_rows:
            topic_order.append(topic_name)
            topic_rows[topic_name] = []

        topic_rows[topic_name].append({
            "row_num": row_idx,
            "platform": platform,
            "content": content or "",
            "status": status or "Draft",
            "tweet_url": tweet_url or "",
        })

    wb.close()

    if topic_index >= len(topic_order):
        raise ValueError(
            f"topic_index {topic_index} out of range (0-{len(topic_order)-1}). "
            f"Workbook has {len(topic_order)} topics."
        )

    topic_name = topic_order[topic_index]
    twitter_rows = [r for r in topic_rows[topic_name] if "twitter" in r["platform"]]
    return twitter_rows, topic_name


def update_excel_after_publish(excel_path: Path, row_nums: list, tweet_urls: list):
    """
    Set Status (col O) = "Published" and Tweet_URL (col P) = comma-joined URLs.
    Creates Tweet_URL column header if absent (backward-compatible).
    Emits TWEET_URLS:[...] to stdout for Flask endpoint to parse.
    """
    wb = openpyxl.load_workbook(str(excel_path))
    ws = wb["Content"]

    # Ensure Tweet_URL header in col P (column 16, 1-based)
    if ws.cell(row=1, column=16).value is None:
        ws.cell(row=1, column=16).value = "Tweet_URL"

    tweet_urls_str = ", ".join(tweet_urls)
    for row_num in row_nums:
        ws.cell(row=row_num, column=15).value = "Published"
        ws.cell(row=row_num, column=16).value = tweet_urls_str

    wb.save(str(excel_path))
    wb.close()
    print(f"TWEET_URLS:{json.dumps(tweet_urls)}")


def find_excel(client_id: str, week_of: str) -> Path:
    """Find the weekly Excel workbook for the given client and week."""
    output_dir = client_config.get_output_dir(client_id)
    normal = output_dir / f"{week_of}-weekly-content.xlsx"
    mock = output_dir / f"{week_of}-mock-weekly-content.xlsx"
    if normal.exists():
        return normal
    if mock.exists():
        return mock
    raise FileNotFoundError(
        f"No workbook found for client={client_id}, week={week_of}.\n"
        f"Expected: {normal}\nRun the pipeline first."
    )


def main():
    parser = argparse.ArgumentParser(description="Publish a Twitter thread to X")
    parser.add_argument("--client", default=None, help="Client ID")
    parser.add_argument("--week-of", required=True, help="Week date YYYY-MM-DD")
    parser.add_argument("--topic-index", type=int, required=True, help="Topic index (0-based)")
    parser.add_argument("--mock", action="store_true", help="Dry run, no API calls")
    parser.add_argument("--lang", default="en", choices=["en"], help="Language (en only)")
    args = parser.parse_args()

    client_id = args.client or client_config.get_active_client()

    # Check x_publishing enabled
    if not client_config.is_x_publishing_enabled(client_id):
        print(f"X publishing is not enabled for client '{client_id}'.")
        print(f"Set x_publishing.enabled = true in clients/{client_id}/config.json")
        sys.exit(0)

    # Find Excel workbook
    excel_path = find_excel(client_id, args.week_of)
    print(f"Workbook: {excel_path}")

    # Read twitter rows for this topic
    twitter_rows, topic_name = read_twitter_rows(excel_path, args.topic_index)

    if not twitter_rows:
        print(f"No Twitter content found for topic_index={args.topic_index}")
        sys.exit(0)

    # Duplicate prevention: already published
    first_row = twitter_rows[0]
    if first_row["status"] == "Published":
        print(f"Already published: topic '{topic_name}' (status=Published)")
        tweet_url = first_row.get("tweet_url", "")
        if tweet_url:
            urls = [u.strip() for u in tweet_url.split(",") if u.strip()]
            print(f"TWEET_URLS:{json.dumps(urls)}")
        sys.exit(0)

    content = first_row["content"]
    if not content:
        print(f"No Twitter content found for topic '{topic_name}'")
        sys.exit(1)

    print(f"Publishing topic [{args.topic_index}]: {topic_name}")
    tweets = split_thread(content)
    print(f"Thread: {len(tweets)} tweet(s)")

    if not args.mock:
        x_client = load_x_client(client_id)
    else:
        x_client = None

    tweet_urls = post_thread(x_client, tweets, mock=args.mock)

    if not args.mock:
        row_nums = [r["row_num"] for r in twitter_rows]
        update_excel_after_publish(excel_path, row_nums, tweet_urls)
    else:
        print(f"TWEET_URLS:{json.dumps(tweet_urls)}")


if __name__ == "__main__":
    main()
