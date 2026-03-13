#!/usr/bin/env python3
"""
Airtable Sync

Pushes weekly content from the local Excel workbook to the client's Airtable base.
Each week gets its own Airtable table: "Week-YYYY-MM-DD".
Records map 1:1 to Excel rows (15 content fields + Week + Client = 17 fields total).

Image attachments: Airtable's API accepts images via public URLs (not local file upload).
Images must be deployed to a public host (e.g. Cloudflare Pages via /deploy) first.
Set airtable.images_base_url in config.json to the deployed site's root URL.
If images_base_url is empty, image fields are stored as plain text paths instead.

Airtable is opt-in per client — controlled by airtable.enabled in config.json.
If not enabled, this script exits cleanly with a skip message.

Usage:
  python scripts/airtable_sync.py --week-of 2026-02-23
  python scripts/airtable_sync.py --week-of 2026-02-23 --client bobe
  python scripts/airtable_sync.py --week-of 2026-02-23 --mock
  python scripts/airtable_sync.py --week-of 2026-02-23 --skip-images

Environment (in .env):
  AIRTABLE_API_KEY — Personal Access Token from airtable.com/create/tokens

Config (clients/{client_id}/config.json):
  airtable.enabled         — true/false
  airtable.base_id         — Airtable Base ID (starts with "app")
  airtable.api_key_env     — env var name for the API key (default: AIRTABLE_API_KEY)
  airtable.images_base_url — deployed site URL for image CDN (e.g. https://bobe.pages.dev)
                             Set after running /deploy. Leave empty to store paths as text.

Setup:
  See reference/airtable-client-setup.md for step-by-step instructions.
"""

import os
import sys
import time
import argparse
from pathlib import Path

import requests
from dotenv import load_dotenv

load_dotenv()

sys.path.insert(0, str(Path(__file__).parent))
import client_config

PROJECT_ROOT = Path(__file__).parent.parent

AIRTABLE_API_BASE = "https://api.airtable.com/v0"
AIRTABLE_META_BASE = "https://api.airtable.com/v0/meta/bases"

# Airtable batch limit: max 10 records per POST/PATCH
BATCH_SIZE = 10

# Fields that hold images — stored as multipleAttachments when images_base_url is set,
# or as singleLineText (path strings) when it is not.
ATTACHMENT_FIELD_EN = "Image_Path"
ATTACHMENT_FIELD_RU = "Image_Path_RU"

# Excel column index → Airtable field name (1-based, 15-column workbook schema)
# Column B (index 2) is now "Bucket" — all subsequent columns shifted +1 vs old schema
COLUMN_MAP = {
    1: "Date",
    2: "Bucket",
    3: "Day",
    4: "Topic",
    5: "Platform",
    6: "Format",
    7: "Content",
    8: "Image_Prompt",
    9: "Image_Path",
    10: "Hashtags",
    11: "Content_RU",
    12: "Image_Prompt_RU",
    13: "Image_Path_RU",
    14: "Hashtags_RU",
    15: "Status",
    16: "Tweet_URL",
}

# Expected column headers for validation (header text -> 1-based column index)
EXPECTED_HEADERS = {v: k for k, v in COLUMN_MAP.items()}


def build_table_fields(use_attachments: bool) -> list:
    """
    Return the TABLE_FIELDS list for table creation.
    Image fields are multipleAttachments when images_base_url is configured,
    singleLineText otherwise.
    """
    image_type = "multipleAttachments" if use_attachments else "singleLineText"
    return [
        {"name": "Date", "type": "singleLineText"},
        {"name": "Bucket", "type": "singleLineText"},
        {"name": "Day", "type": "singleLineText"},
        {"name": "Topic", "type": "multilineText"},
        {"name": "Platform", "type": "singleLineText"},
        {"name": "Format", "type": "singleLineText"},
        {"name": "Content", "type": "multilineText"},
        {"name": "Image_Prompt", "type": "multilineText"},
        {"name": "Image_Path", "type": image_type},
        {"name": "Hashtags", "type": "singleLineText"},
        {"name": "Content_RU", "type": "multilineText"},
        {"name": "Image_Prompt_RU", "type": "multilineText"},
        {"name": "Image_Path_RU", "type": image_type},
        {"name": "Hashtags_RU", "type": "singleLineText"},
        {"name": "Status", "type": "singleLineText"},
        {"name": "Week", "type": "singleLineText"},
        {"name": "Client", "type": "singleLineText"},
    ]


def get_api_key(at_config: dict) -> str:
    """Read the Airtable API key from environment."""
    env_var = at_config.get("api_key_env", "AIRTABLE_API_KEY")
    key = os.environ.get(env_var)
    if not key:
        print(f"Error: {env_var} not set in .env")
        print("See reference/airtable-client-setup.md for setup instructions.")
        sys.exit(1)
    return key


def get_headers(api_key: str) -> dict:
    return {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }


def image_path_to_url(path_str: str, images_base_url: str) -> str | None:
    """
    Convert a local image path from the workbook to a public CDN URL.

    The workbook stores paths like:
      outputs/content/bobe/images/2026-02-16-weekly/filename.png

    The deployed site serves images at:
      {images_base_url}/images/2026-02-16-weekly/filename.png

    So we strip the leading 'outputs/content/{client_id}/' prefix.
    Returns None if path_str is empty.
    """
    if not path_str or not path_str.strip():
        return None

    p = Path(path_str)

    # The images/ segment marks where the public path starts
    try:
        # Find 'images' in the path parts and take from there
        parts = p.parts
        images_idx = next(i for i, part in enumerate(parts) if part == "images")
        relative = "/".join(parts[images_idx:])
    except StopIteration:
        # Fallback: just use the filename
        relative = p.name

    base = images_base_url.rstrip("/")
    return f"{base}/{relative}"


def get_or_create_table(base_id: str, week_of: str, headers: dict,
                        use_attachments: bool, mock: bool = False) -> str:
    """
    Check if table 'Week-{week_of}' exists in the base. Create it if not.
    Returns the table ID.
    """
    table_name = f"Week-{week_of}"

    if mock:
        mode = "attachment" if use_attachments else "text"
        print(f"  [MOCK] Would check/create table '{table_name}' (image mode: {mode})")
        return "tblMOCK000000"

    # List existing tables
    url = f"{AIRTABLE_META_BASE}/{base_id}/tables"
    resp = requests.get(url, headers=headers, timeout=30)

    if resp.status_code == 200:
        tables = resp.json().get("tables", [])
        for table in tables:
            if table["name"] == table_name:
                print(f"  Table '{table_name}' already exists (ID: {table['id']})")
                return table["id"]
    elif resp.status_code == 403:
        print("Error: Airtable token missing 'schema.bases:read' scope.")
        print("See reference/airtable-client-setup.md — Token Scopes section.")
        sys.exit(1)
    else:
        print(f"Warning: Could not list tables (HTTP {resp.status_code}). Attempting to create.")

    # Create the table
    mode = "image attachments" if use_attachments else "text paths"
    print(f"  Creating table '{table_name}' ({mode})...")
    fields = build_table_fields(use_attachments)
    resp = requests.post(
        f"{AIRTABLE_META_BASE}/{base_id}/tables",
        headers=headers,
        json={"name": table_name, "fields": fields},
        timeout=30,
    )

    if resp.status_code in (200, 201):
        table_id = resp.json()["id"]
        print(f"  Table created: '{table_name}' (ID: {table_id})")
        return table_id
    else:
        print(f"Error: Failed to create table '{table_name}': HTTP {resp.status_code}")
        print(resp.text[:500])
        sys.exit(1)


def read_excel_rows(excel_path: Path, week_of: str, client_id: str) -> list:
    """
    Read all content rows from the Excel workbook's Content sheet.
    Returns list of dicts with Airtable field names as keys.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    if not excel_path.exists():
        print(f"Error: Workbook not found: {excel_path}")
        sys.exit(1)

    wb = load_workbook(str(excel_path), read_only=True, data_only=True)

    if "Content" not in wb.sheetnames:
        print(f"Error: 'Content' sheet not found in {excel_path}")
        sys.exit(1)

    ws = wb["Content"]

    # Validate column headers (row 1)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row:
        col_map = {}
        for idx, val in enumerate(header_row, 1):
            if val and str(val).strip() in EXPECTED_HEADERS:
                col_map[str(val).strip()] = idx
        missing = [h for h in ["Date", "Topic", "Platform", "Content", "Status"] if h not in col_map]
        if missing:
            print(f"Warning: Excel headers missing required columns: {', '.join(missing)}")
            print(f"  Found headers: {[str(h) for h in header_row if h]}")
            print(f"  Falling back to positional column mapping")
            col_map = None
    else:
        col_map = None

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell for cell in row):
            continue

        record = {}
        for col_idx, field_name in COLUMN_MAP.items():
            if col_map and field_name in col_map:
                actual_idx = col_map[field_name] - 1
            else:
                actual_idx = col_idx - 1
            val = row[actual_idx] if actual_idx < len(row) else None
            record[field_name] = str(val) if val is not None else ""

        record["Week"] = week_of
        record["Client"] = client_id
        rows.append(record)

    wb.close()
    print(f"  Read {len(rows)} rows from {excel_path.name}")
    return rows


def build_record_fields(row: dict, images_base_url: str, skip_images: bool) -> dict:
    """
    Build the Airtable fields dict for one row.

    When images_base_url is set and skip_images is False:
      - Image_Path and Image_Path_RU become multipleAttachments arrays with {url, filename}
    Otherwise:
      - Image_Path and Image_Path_RU remain plain text strings
    """
    fields = {}
    use_attachments = bool(images_base_url) and not skip_images

    for field_name, value in row.items():
        if field_name in (ATTACHMENT_FIELD_EN, ATTACHMENT_FIELD_RU):
            if use_attachments:
                url = image_path_to_url(value, images_base_url)
                if url:
                    filename = Path(value).name if value else "image.png"
                    fields[field_name] = [{"url": url, "filename": filename}]
                # If no URL, omit the field entirely (leave attachment empty)
            else:
                # Store as plain text path
                fields[field_name] = value
        else:
            fields[field_name] = value

    return fields


def push_records(base_id: str, table_id: str, rows: list, headers: dict,
                 images_base_url: str, skip_images: bool, mock: bool = False) -> int:
    """
    Create all records in Airtable, including image attachments if configured.
    Returns the count of successfully pushed records.
    """
    use_attachments = bool(images_base_url) and not skip_images

    if mock:
        print(f"  [MOCK] Would push {len(rows)} records (images: {'as attachments' if use_attachments else 'as text'})")
        for i, row in enumerate(rows[:3], 1):
            en_url = image_path_to_url(row.get("Image_Path", ""), images_base_url) if use_attachments else None
            print(f"    Record {i}: {row.get('Date', '?')} | {row.get('Platform', '?')} | {row.get('Topic', '?')[:40]}")
            if en_url:
                print(f"      EN image: {en_url}")
        if len(rows) > 3:
            print(f"    ... and {len(rows) - 3} more")
        return len(rows)

    url = f"{AIRTABLE_API_BASE}/{base_id}/{table_id}"
    pushed = 0

    for batch_start in range(0, len(rows), BATCH_SIZE):
        batch = rows[batch_start:batch_start + BATCH_SIZE]
        payload = {
            "records": [
                {"fields": build_record_fields(row, images_base_url, skip_images)}
                for row in batch
            ]
        }

        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=30)

            if resp.status_code in (200, 201):
                pushed += len(resp.json().get("records", []))
                print(f"  Pushed batch {batch_start // BATCH_SIZE + 1}: {len(batch)} records (total: {pushed})")
            else:
                print(f"  Warning: Batch {batch_start // BATCH_SIZE + 1} failed: HTTP {resp.status_code}")
                print(f"  {resp.text[:400]}")

        except Exception as e:
            print(f"  Warning: Batch {batch_start // BATCH_SIZE + 1} error: {e}")

        time.sleep(0.25)  # Stay within 5 req/sec rate limit

    return pushed


def main():
    parser = argparse.ArgumentParser(
        description="Push weekly content from Excel workbook to Airtable with image attachments."
    )
    parser.add_argument("--week-of", required=True,
                        help="Monday of target week (YYYY-MM-DD)")
    parser.add_argument("--mock", action="store_true",
                        help="Dry run — print what would happen without making API calls")
    parser.add_argument("--skip-images", action="store_true",
                        help="Push text content only, store image paths as plain text")
    parser.add_argument("--images-base-url", default=None,
                        help="Override the deployed site URL from config (e.g. https://bobe.pages.dev)")
    client_config.add_client_arg(parser)

    args = parser.parse_args()
    active_client = client_config.resolve_client(args)
    week_of = args.week_of

    # Load client config
    config = client_config.load_config(active_client)
    display_name = config.get("display_name", active_client)
    at_config = client_config.get_airtable_config(active_client)

    # Resolve images_base_url: CLI flag > config > empty
    images_base_url = (
        args.images_base_url
        or at_config.get("images_base_url", "").strip()
    )

    print(f"\nAirtable Sync — {display_name} ({active_client})")
    print(f"Week: {week_of}")

    if args.skip_images:
        print(f"Images: skipped (--skip-images)")
    elif images_base_url:
        print(f"Images: attached from {images_base_url}")
    else:
        print(f"Images: stored as text paths (set airtable.images_base_url in config or run /deploy first)")

    # Check if Airtable is enabled
    if not client_config.is_airtable_enabled(active_client):
        if not at_config.get("enabled"):
            print(f"\nAirtable not enabled for {active_client} — skipping.")
            print("To enable: set airtable.enabled = true and airtable.base_id in config.json")
        elif not at_config.get("base_id"):
            print(f"\nAirtable enabled but base_id is empty — skipping.")
            print(f"To fix: set airtable.base_id in clients/{active_client}/config.json")
        sys.exit(0)

    base_id = at_config["base_id"]
    print(f"Base ID: {base_id}")

    # Get API key and headers
    api_key = get_api_key(at_config)
    headers = get_headers(api_key)

    # Locate the Excel workbook
    output_dir = client_config.get_output_dir(active_client)
    excel_path = output_dir / f"{week_of}-weekly-content.xlsx"

    # Read rows
    print(f"\nReading workbook: {excel_path}")
    rows = read_excel_rows(excel_path, week_of, active_client)

    if not rows:
        print("No rows found in workbook. Nothing to push.")
        sys.exit(0)

    # Count how many images we'll attach
    use_attachments = bool(images_base_url) and not args.skip_images
    if use_attachments:
        en_count = sum(1 for r in rows if image_path_to_url(r.get("Image_Path", ""), images_base_url))
        ru_count = sum(1 for r in rows if image_path_to_url(r.get("Image_Path_RU", ""), images_base_url))
        print(f"  Will attach: {en_count} EN images + {ru_count} RU images")

    # Get or create the Airtable table
    print(f"\nChecking Airtable table...")
    table_id = get_or_create_table(
        base_id, week_of, headers,
        use_attachments=use_attachments,
        mock=args.mock,
    )

    # Push all records (with images inline if configured)
    print(f"\nPushing {len(rows)} records...")
    pushed = push_records(
        base_id, table_id, rows, headers,
        images_base_url=images_base_url,
        skip_images=args.skip_images,
        mock=args.mock,
    )

    # Summary
    print(f"\n{'='*60}")
    if args.mock:
        print(f"  [MOCK] Airtable Sync — {display_name}")
        print(f"  Would push: {pushed} records")
        if use_attachments:
            print(f"  Images from: {images_base_url}")
        print(f"  Base:  {base_id}")
        print(f"  Table: Week-{week_of}")
    else:
        print(f"  Airtable Sync Complete — {display_name}")
        print(f"  Records pushed: {pushed}/{len(rows)}")
        if use_attachments:
            print(f"  Images attached from: {images_base_url}")
        elif not args.skip_images:
            print(f"  Images: stored as text paths")
            print(f"  Tip: run /deploy then set airtable.images_base_url to enable inline images")
        print(f"  Base:  {base_id}")
        print(f"  Table: Week-{week_of}")
        print(f"  View:  https://airtable.com/{base_id}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
